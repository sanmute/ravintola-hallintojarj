"""
poweresta_exporter.py — Vie reseptejä PoweResta-Excel-muotoon (yksi resepti
per välilehti), samaa mallia kuin PoweResta-tuonnin odottama rakenne.

Määrät kirjoitetaan kiloina. Excelin lukuformaatti '0.000' käyttää pistettä
vain FORMAATTIKOODIN syntaksissa — Excel näyttää sen käyttäjän Windows/Office
-asetusten mukaisella desimaalierottimella (pilkulla suomenkielisessä
asennuksessa), joten arvoa EI pidä yrittää kirjoittaa merkkijonona pilkulla.
"""

import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Real units used throughout this app (see recipe_scraper_v2.structure_ingredients
# and the "Oma resepti" contribute form's ING_UNITS) — not the generic
# English tsp/tbsp/cup set, which never actually appears in this database.
_KG_PER_UNIT = {
    'kg': 1.0, 'g': 0.001, 'mg': 0.000001,
    'l': 1.0, 'dl': 0.1, 'cl': 0.01, 'ml': 0.001,
    'tl': 0.005,    # teelusikka ≈ 5 g
    'rkl': 0.015,   # ruokalusikka ≈ 15 g
    'kpl': 0.1, 'pkt': 0.1, 'pss': 0.1, 'prk': 0.2,
    'rs': 0.1, 'purkki': 0.2, 'rasia': 0.2, 'nippu': 0.1, 'ripaus': 0.001,
}

_INVALID_SHEET_CHARS = re.compile(r'[\\/*?:\[\]]')


class PoweRestaExporter:
    """Kokoaa valitut reseptit yhdeksi PoweResta-muotoiseksi työkirjaksi."""

    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self._used_sheet_names = set()

    def _unique_sheet_name(self, name_fi):
        base = _INVALID_SHEET_CHARS.sub(' ', name_fi).strip() or 'Resepti'
        base = base[:31]
        candidate = base
        n = 2
        while candidate in self._used_sheet_names:
            suffix = f' ({n})'
            candidate = base[:31 - len(suffix)] + suffix
            n += 1
        self._used_sheet_names.add(candidate)
        return candidate

    def _to_kg(self, quantity, unit):
        if quantity is None:
            return None
        multiplier = _KG_PER_UNIT.get((unit or '').strip().lower())
        if multiplier is None:
            return None  # tuntematon yksikkö — ei arvata väärin, jätetään tyhjäksi
        return quantity * multiplier

    def add_recipe(self, name_fi, ingredients, servings=None, diets=None):
        """ingredients: list of {'name': str, 'quantity': float|None, 'unit': str|None}"""
        ws = self.wb.create_sheet(title=self._unique_sheet_name(name_fi))

        ws['A2'] = 'Nimi'
        ws['B2'] = name_fi
        ws['A3'] = 'Reseptiryhmät'
        ws['B3'] = ''
        ws['A4'] = 'Annoskoko (g)'
        ws['B4'] = ''  # ei tiedossa mistään lähteestä — täytetään PoweRestassa
        ws['A5'] = 'Annosmäärä'
        ws['B5'] = servings if servings else ''
        ws['A6'] = 'Ruokavaliot'
        ws['B6'] = diets or ''
        ws['A7'] = 'Kypsymishävikki'
        ws['B7'] = 0.05
        ws['A8'] = 'Jakeluhävikki'
        ws['B8'] = 0

        headers = ['Rivit', 'Otsikko', 'Raaka-aine / työohje', 'Valmistamaton paino',
                   'Esikäsittelyhävikki', 'Ostopaino', 'Onko aliresepti']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=12, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')

        ws['B13'] = 'Valmistus'
        ws['B13'].font = Font(bold=True)

        row = 14
        for ing in ingredients:
            kg = self._to_kg(ing.get('quantity'), ing.get('unit'))
            ws.cell(row=row, column=3, value=ing.get('name') or '')
            if kg is not None:
                c = ws.cell(row=row, column=4, value=round(kg, 4))
                c.number_format = '0.000'
                f = ws.cell(row=row, column=6, value=round(kg, 4))
                f.number_format = '0.000'
            ws.cell(row=row, column=5, value=0)
            row += 1

        widths = {'A': 25, 'B': 30, 'C': 35, 'D': 15, 'E': 15, 'F': 15, 'G': 15}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

    def get_bytes(self):
        out = BytesIO()
        self.wb.save(out)
        out.seek(0)
        return out

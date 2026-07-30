"""
order_catalog.py — Viikoittainen tilausluettelo (tuotekatalogi).

Kokoaa viikon ruokalistan reseptien raaka-aineet yhteen tilauslistaksi,
jonka määriä käyttäjä voi säätää kysynnän ja varastotilanteen mukaan.
Säädöt säilyvät, vaikka lista laskettaisiin uudelleen (esim. ruokalistan
muututtua): laskennallinen määrä päivittyy, käsin asetettu määrä pysyy.

Käyttöönotto app.py:ssä (initien jälkeen):

    from order_catalog import init_order_catalog
    init_order_catalog(app, DB_PATH)

Reitit:
  POST   /api/catalog/generate           {plan_id, week}  luo/päivitä
  GET    /api/catalog?plan_id=&week=                      hae rivit
  PATCH  /api/catalog/item/<id>          {adjusted_qty|note|checked}
  POST   /api/catalog/item               {catalog_id, name, unit, qty}
  DELETE /api/catalog/item/<id>
  GET    /api/catalog/export?plan_id=&week=               Excel
"""

import io
import os
import sqlite3
from datetime import datetime, timedelta

from flask import jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont


def _conn(db_path):
    conn = sqlite3.connect(db_path, timeout=10)
    conn.row_factory = sqlite3.Row
    return conn




def _ensure_tables(db_path):
    conn = _conn(db_path)
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS order_catalogs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            meal_plan_id INTEGER NOT NULL,
            week_number INTEGER NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now')),
            UNIQUE (meal_plan_id, week_number)
        );
        CREATE TABLE IF NOT EXISTS order_catalog_items (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            catalog_id INTEGER NOT NULL REFERENCES order_catalogs(id)
                ON DELETE CASCADE,
            name TEXT NOT NULL,
            unit TEXT DEFAULT '',
            calculated_qty REAL,          -- resepteistä laskettu
            adjusted_qty REAL,            -- käyttäjän säätämä (NULL = käytä laskettua)
            note TEXT DEFAULT '',
            checked INTEGER DEFAULT 0,    -- 'kerätty/tilattu' -rasti
            manual INTEGER DEFAULT 0,     -- käsin lisätty rivi
            UNIQUE (catalog_id, name, unit)
        );
    """)
    cols = [r[1] for r in conn.execute('PRAGMA table_info(order_catalogs)').fetchall()]
    if 'target_servings' not in cols:
        # Montako annosta viikon tilaus mitoitetaan — ravintola tarjoilee
        # yleensä 125-150 annosta/päivä, mutta reseptit on kirjoitettu
        # vaihtelevalle annosmäärälle (4-10+), joten raaka-aineet pitää
        # skaalata per resepti sen omalla annosmäärällä ennen summausta.
        conn.execute('ALTER TABLE order_catalogs ADD COLUMN target_servings INTEGER')
    conn.commit()
    conn.close()


def _effective_recipe_ids(conn, plan_id, week):
    """Recipe ids actually served that week, mirror-aware: Hoiva's ma-pe
    always equals its linked Kesti plan (meal_plans.mirrors_plan_id), only
    la-su is Hoiva's own data. Duplicates are kept (the same recipe served
    on two different days that week must be counted/ordered twice)."""
    plan = conn.execute('SELECT mirrors_plan_id FROM meal_plans WHERE id = ?', (plan_id,)).fetchone()
    mirrors = plan['mirrors_plan_id'] if plan else None
    own = conn.execute(
        'SELECT day_of_week, recipe_id FROM meal_plan_days WHERE meal_plan_id = ? AND week_number = ?',
        (plan_id, week)).fetchall()
    if not mirrors:
        return [r['recipe_id'] for r in own]
    mirrored = conn.execute(
        'SELECT recipe_id FROM meal_plan_days WHERE meal_plan_id = ? AND week_number = ? AND day_of_week < 5',
        (mirrors, week)).fetchall()
    return [r['recipe_id'] for r in mirrored] + [r['recipe_id'] for r in own if r['day_of_week'] >= 5]


def _aggregate_ingredients(conn, plan_id, week, target_servings=None):
    """Sum ingredient quantities for one week of a meal plan.

    target_servings: if given, each recipe's ingredient quantities are
    scaled by target_servings / recipe's own servings count — recipes are
    written for whatever batch they were originally portioned for (commonly
    4-10, but some imported recipes are already bulk/production-batch sized
    with no servings count saved), and the kitchen needs one consistent
    total for however many people are actually being served that week.
    Recipes with no servings value are left unscaled (factor 1) rather than
    guessed — assuming a small default like 4 for what's often already a
    bulk recipe produced wildly inflated orders in practice.

    Returns (rows, missing_servings) — missing_servings is the sorted list
    of distinct recipe names that had no servings value and so were left
    unscaled.
    """
    recipe_ids = _effective_recipe_ids(conn, plan_id, week)
    if not recipe_ids:
        return [], []

    distinct_ids = list(set(recipe_ids))
    placeholders = ','.join('?' * len(distinct_ids))
    recipe_rows = conn.execute(
        f'SELECT id, name_fi, servings FROM recipes WHERE id IN ({placeholders})',
        distinct_ids).fetchall()
    servings_by_recipe = {r['id']: r['servings'] for r in recipe_rows}
    name_by_recipe = {r['id']: r['name_fi'] for r in recipe_rows}
    missing_servings = sorted({name_by_recipe[rid] for rid in distinct_ids
                              if not servings_by_recipe.get(rid)}) if target_servings else []

    ing_rows = conn.execute(
        f'''SELECT ri.recipe_id, i.name_fi AS name, ri.unit AS unit, ri.quantity AS quantity
            FROM recipe_ingredients ri JOIN ingredients i ON i.id = ri.ingredient_id
            WHERE ri.recipe_id IN ({placeholders})''', distinct_ids).fetchall()
    ing_by_recipe = {}
    for row in ing_rows:
        ing_by_recipe.setdefault(row['recipe_id'], []).append(row)

    sums = {}     # key -> running total (only from occurrences with a known quantity)
    has_known = set()
    for rid in recipe_ids:
        base = servings_by_recipe.get(rid)
        # No guessed default here on purpose: many recipes without a saved
        # servings count turned out to already be bulk/production-batch
        # recipes (kg-scale ingredients), not small per-4-people ones —
        # assuming 4 and scaling up produced wildly inflated orders (one
        # real case: 7kg lettuce -> 245kg). Safer to leave them unscaled
        # and flag them, than to guess and risk a 30x+ over-order.
        factor = (target_servings / base) if (target_servings and base) else 1
        for ing in ing_by_recipe.get(rid, []):
            key = (ing['name'], ing['unit'] or '')
            if ing['quantity'] is None:
                sums.setdefault(key, 0)
                continue
            sums[key] = sums.get(key, 0) + ing['quantity'] * factor
            has_known.add(key)

    rows = [{'name': k[0], 'unit': k[1], 'total': (round(v, 3) if k in has_known else None)}
           for k, v in sorted(sums.items())]
    return rows, missing_servings


def init_order_catalog(app, db_path):
    _ensure_tables(db_path)

    # ---------------------------------------------------- generate / refresh
    @app.route('/api/catalog/generate', methods=['POST'])
    def catalog_generate():
        d = request.get_json(force=True)
        plan_id, week = d.get('plan_id'), d.get('week')
        if not plan_id or not week:
            return jsonify({'error': 'plan_id ja week vaaditaan'}), 400
        target_servings = d.get('target_servings')
        if target_servings is not None:
            try:
                target_servings = int(target_servings)
                if not (1 <= target_servings <= 2000):
                    return jsonify({'error': 'Annosmäärän pitää olla 1-2000'}), 400
            except (TypeError, ValueError):
                return jsonify({'error': 'Virheellinen annosmäärä'}), 400

        conn = _conn(db_path)
        try:
            ings, missing_servings = _aggregate_ingredients(conn, plan_id, week, target_servings)

            cat = conn.execute(
                'SELECT id FROM order_catalogs WHERE meal_plan_id=? AND week_number=?',
                (plan_id, week)).fetchone()
            if cat:
                cat_id = cat['id']
            else:
                cur = conn.execute(
                    'INSERT INTO order_catalogs (meal_plan_id, week_number) VALUES (?,?)',
                    (plan_id, week))
                cat_id = cur.lastrowid

            # Upsert calculated rows; user adjustments (adjusted_qty, note,
            # checked) are preserved because we only touch calculated_qty.
            new_names = set()
            for ing in ings:
                new_names.add((ing['name'], ing['unit'] or ''))
                conn.execute(
                    '''INSERT INTO order_catalog_items
                           (catalog_id, name, unit, calculated_qty)
                       VALUES (?,?,?,?)
                       ON CONFLICT(catalog_id, name, unit)
                       DO UPDATE SET calculated_qty = excluded.calculated_qty''',
                    (cat_id, ing['name'], ing['unit'] or '', ing['total']))

            # Remove auto rows whose ingredient no longer appears (keep
            # manual rows and rows the user adjusted or annotated).
            for row in conn.execute(
                    '''SELECT id, name, unit FROM order_catalog_items
                       WHERE catalog_id=? AND manual=0 AND adjusted_qty IS NULL
                         AND (note IS NULL OR note='') AND checked=0''',
                    (cat_id,)).fetchall():
                if (row['name'], row['unit'] or '') not in new_names:
                    conn.execute('DELETE FROM order_catalog_items WHERE id=?',
                                 (row['id'],))

            conn.execute(
                "UPDATE order_catalogs SET updated_at=datetime('now'), target_servings=? WHERE id=?",
                (target_servings, cat_id))
            conn.commit()
            n = conn.execute(
                'SELECT COUNT(*) c FROM order_catalog_items WHERE catalog_id=?',
                (cat_id,)).fetchone()['c']
        finally:
            conn.close()
        if not ings and n == 0:
            return jsonify({'catalog_id': cat_id, 'items': n,
                            'warning': 'Viikon resepteillä ei ole raaka-ainetietoja. '
                                       'Voit lisätä rivejä käsin.'})
        resp = {'catalog_id': cat_id, 'items': n}
        if missing_servings:
            resp['missing_servings'] = missing_servings
            resp['warning'] = (
                f"{len(missing_servings)} reseptillä ei ole annosmäärää tallennettuna — "
                f"niiden raaka-aineita EI skaalattu (määrät ovat reseptin alkuperäisiä, "
                f"tarkista tarvittaessa käsin): "
                + ', '.join(missing_servings[:8]) + ('…' if len(missing_servings) > 8 else ''))
        return jsonify(resp)

    # ---------------------------------------------------- read
    @app.route('/api/catalog')
    def catalog_get():
        plan_id = request.args.get('plan_id', type=int)
        week = request.args.get('week', type=int)
        conn = _conn(db_path)
        try:
            cat = conn.execute(
                'SELECT * FROM order_catalogs WHERE meal_plan_id=? AND week_number=?',
                (plan_id, week)).fetchone()
            if not cat:
                return jsonify({'catalog': None, 'items': []})
            items = conn.execute(
                '''SELECT * FROM order_catalog_items WHERE catalog_id=?
                   ORDER BY checked, name COLLATE NOCASE''',
                (cat['id'],)).fetchall()
        finally:
            conn.close()
        return jsonify({'catalog': dict(cat),
                        'items': [dict(i) for i in items]})

    # ---------------------------------------------------- edit one row
    @app.route('/api/catalog/item/<int:item_id>', methods=['PATCH'])
    def catalog_item_patch(item_id):
        d = request.get_json(force=True)
        sets, vals = [], []
        if 'adjusted_qty' in d:
            v = d['adjusted_qty']
            if v is not None:
                try:
                    v = round(float(v), 3)
                    if v < 0:
                        return jsonify({'error': 'Määrä ei voi olla negatiivinen'}), 400
                except (TypeError, ValueError):
                    return jsonify({'error': 'Virheellinen määrä'}), 400
            sets.append('adjusted_qty=?'); vals.append(v)
        if 'note' in d:
            sets.append('note=?'); vals.append(str(d['note'])[:200])
        if 'checked' in d:
            sets.append('checked=?'); vals.append(1 if d['checked'] else 0)
        if not sets:
            return jsonify({'error': 'Ei muutettavia kenttiä'}), 400
        vals.append(item_id)
        conn = _conn(db_path)
        try:
            cur = conn.execute(
                f"UPDATE order_catalog_items SET {', '.join(sets)} WHERE id=?", vals)
            conn.commit()
            if cur.rowcount == 0:
                return jsonify({'error': 'Riviä ei löydy'}), 404
            row = conn.execute('SELECT * FROM order_catalog_items WHERE id=?',
                               (item_id,)).fetchone()
        finally:
            conn.close()
        return jsonify(dict(row))

    # ---------------------------------------------------- manual add / delete
    @app.route('/api/catalog/item', methods=['POST'])
    def catalog_item_add():
        d = request.get_json(force=True)
        name = (d.get('name') or '').strip()
        if not name or not d.get('catalog_id'):
            return jsonify({'error': 'catalog_id ja nimi vaaditaan'}), 400
        try:
            qty = round(float(d.get('qty', 0)), 3)
        except (TypeError, ValueError):
            return jsonify({'error': 'Virheellinen määrä'}), 400
        conn = _conn(db_path)
        try:
            cur = conn.execute(
                '''INSERT INTO order_catalog_items
                       (catalog_id, name, unit, adjusted_qty, manual)
                   VALUES (?,?,?,?,1)''',
                (d['catalog_id'], name, (d.get('unit') or '').strip(), qty))
            conn.commit()
            row = conn.execute('SELECT * FROM order_catalog_items WHERE id=?',
                               (cur.lastrowid,)).fetchone()
        except sqlite3.IntegrityError:
            return jsonify({'error': 'Tuote on jo listalla'}), 409
        finally:
            conn.close()
        return jsonify(dict(row))

    @app.route('/api/catalog/item/<int:item_id>', methods=['DELETE'])
    def catalog_item_delete(item_id):
        conn = _conn(db_path)
        try:
            conn.execute('DELETE FROM order_catalog_items WHERE id=?', (item_id,))
            conn.commit()
        finally:
            conn.close()
        return jsonify({'ok': True})

    # ---------------------------------------------------- Excel export
    @app.route('/api/catalog/export')
    def catalog_export():
        plan_id = request.args.get('plan_id', type=int)
        week = request.args.get('week', type=int)
        conn = _conn(db_path)
        try:
            cat = conn.execute(
                'SELECT * FROM order_catalogs WHERE meal_plan_id=? AND week_number=?',
                (plan_id, week)).fetchone()
            if not cat:
                return jsonify({'error': 'Tilauslistaa ei ole luotu tälle viikolle'}), 404
            items = conn.execute(
                '''SELECT * FROM order_catalog_items WHERE catalog_id=?
                   ORDER BY name COLLATE NOCASE''', (cat['id'],)).fetchall()
        finally:
            conn.close()

        wb = Workbook()
        ws = wb.active
        ws.title = f'Viikko {week}'
        ws['A1'] = f'TILAUSLISTA — Viikko {week}'
        ws['A1'].font = Font(bold=True, size=13)
        servings_note = f' — mitoitettu {cat["target_servings"]} annokselle' if cat['target_servings'] else ''
        ws['A2'] = f'Luotu: {datetime.now().strftime("%d.%m.%Y %H:%M")}{servings_note}'

        headers = ['Tuote', 'Yksikkö', 'Laskettu määrä', 'Tilattava määrä',
                   'Huomiot', 'Kerätty']
        for col, h in enumerate(headers, 1):
            c = ws.cell(4, col, h)
            c.font = Font(bold=True)

        r = 5
        for it in items:
            qty = it['adjusted_qty'] if it['adjusted_qty'] is not None \
                else it['calculated_qty']
            ws.cell(r, 1, it['name'])
            ws.cell(r, 2, it['unit'])
            ws.cell(r, 3, it['calculated_qty'])
            ws.cell(r, 4, qty)
            ws.cell(r, 5, it['note'] or '')
            ws.cell(r, 6, 'x' if it['checked'] else '')
            if it['adjusted_qty'] is not None:
                ws.cell(r, 4).font = Font(bold=True)  # korostetaan säädetyt
            r += 1

        for col, width in zip('ABCDEF', (34, 10, 15, 16, 30, 9)):
            ws.column_dimensions[col].width = width

        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        buf.seek(0)
        return send_file(
            buf, as_attachment=True,
            download_name=f'tilauslista_vko{week}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    # ---------------------------------------------------- PDF export (print)
    @app.route('/api/catalog/export/pdf')
    def catalog_export_pdf():
        plan_id = request.args.get('plan_id', type=int)
        week = request.args.get('week', type=int)
        conn = _conn(db_path)
        try:
            cat = conn.execute(
                'SELECT * FROM order_catalogs WHERE meal_plan_id=? AND week_number=?',
                (plan_id, week)).fetchone()
            if not cat:
                return jsonify({'error': 'Tilauslistaa ei ole luotu tälle viikolle'}), 404
            items = conn.execute(
                '''SELECT * FROM order_catalog_items WHERE catalog_id=?
                   ORDER BY name COLLATE NOCASE''', (cat['id'],)).fetchall()
        finally:
            conn.close()

        buf = io.BytesIO()
        doc = SimpleDocTemplate(
            buf, pagesize=A4,
            topMargin=15 * mm, bottomMargin=15 * mm,
            leftMargin=15 * mm, rightMargin=15 * mm,
            title=f'Tilauslista viikko {week}')

        styles = getSampleStyleSheet()
        servings_note = f' — mitoitettu {cat["target_servings"]} annokselle' if cat['target_servings'] else ''
        elements = [
            Paragraph(f'TILAUSLISTA — Viikko {week}', styles['Title']),
            Paragraph(f'Luotu: {datetime.now().strftime("%d.%m.%Y %H:%M")}{servings_note}',
                      styles['Normal']),
            Spacer(1, 8 * mm),
        ]

        # Selkeät solutyylit — Paragraph rivittää pitkän tekstin sarakkeen
        # sisään sen sijaan, että se valuisi viereisen sarakkeen päälle.
        cell_style = ParagraphStyle('cell', parent=styles['Normal'],
                                    fontName='Helvetica', fontSize=9, leading=11)
        cell_style_right = ParagraphStyle('cellRight', parent=cell_style,
                                          alignment=2)  # TA_RIGHT

        header = ['Tuote', 'Yksikkö', 'Laskettu', 'Tilattava', 'Huomiot', 'Kerätty']
        rows = [header]
        for it in items:
            qty = it['adjusted_qty'] if it['adjusted_qty'] is not None \
                else it['calculated_qty']
            rows.append([
                Paragraph(it['name'] or '', cell_style),
                Paragraph(it['unit'] or '', cell_style),
                Paragraph('' if it['calculated_qty'] is None else str(it['calculated_qty']),
                          cell_style_right),
                Paragraph('' if qty is None else str(qty), cell_style_right),
                Paragraph(it['note'] or '', cell_style),
                'x' if it['checked'] else '',
            ])

        # Sarakeleveydet mahtuvat A4:n käytettävään leveyteen
        # (210mm - 15mm*2 marginaalit = 180mm).
        col_widths = [50 * mm, 15 * mm, 20 * mm, 20 * mm, 58 * mm, 15 * mm]
        table = Table(rows, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#5CB85C')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (5, 0), (5, -1), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#cccccc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(table)
        doc.build(elements)
        buf.seek(0)
        return send_file(
            buf, as_attachment=False,
            download_name=f'tilauslista_vko{week}.pdf',
            mimetype='application/pdf')


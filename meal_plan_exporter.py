"""
Excel Export for Meal Plans
Generates Excel files matching the Finnish meal plan template
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from meal_plan_db import MealPlanDB, DishCategory
from collections import defaultdict
import datetime

class MealPlanExporter:
    """Exports meal plans to Excel format"""
    
    # Color mapping for dish categories
    CATEGORY_COLORS = {
        DishCategory.FISH.value: 'ADD8E6',        # Light blue
        DishCategory.CHICKEN.value: 'FFFF99',     # Yellow
        DishCategory.BEEF.value: 'FF6666',        # Red
        DishCategory.VEGETABLE.value: '99CC99',   # Green
        DishCategory.CARB.value: 'FFFFFF'         # White (legend items)
    }
    
    def __init__(self, db_path='meal_plans.db'):
        self.db = MealPlanDB(db_path)
    
    def export_meal_plan(self, meal_plan_id, output_path=None):
        """
        Export a meal plan to Excel
        
        Args:
            meal_plan_id: ID of the meal plan to export
            output_path: Path to save Excel file (default: meal_plan_{id}.xlsx)
        
        Returns:
            Path to the generated file
        """
        
        if not output_path:
            output_path = f'meal_plan_{meal_plan_id}.xlsx'
        
        # Get meal plan data
        meal_plan = self.db.get_meal_plan(meal_plan_id)
        if not meal_plan:
            print(f"❌ Meal plan {meal_plan_id} not found")
            return None
        
        print(f"📊 Exporting meal plan: {meal_plan['name']}")
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Meal Plan"
        
        # Set column widths
        ws.column_dimensions['A'].width = 15
        for col in range(2, 100):  # Dynamic columns for recipes
            ws.column_dimensions[get_column_letter(col)].width = 18
        
        # Create meal plan grid
        self._create_meal_plan_grid(ws, meal_plan)
        
        # Create ingredient legend sheet
        self._create_ingredient_legend(wb, meal_plan)
        
        # Create recipe details sheet
        self._create_recipe_details(wb, meal_plan)
        
        # Save workbook
        wb.save(output_path)
        print(f"✅ Exported to: {output_path}")
        
        return output_path
    
    def _create_meal_plan_grid(self, ws, meal_plan):
        """Create the main meal plan grid (4-week cycles)"""
        
        # Group meals by week
        meals_by_week = defaultdict(list)
        for week_num, day_num, recipe_id, meal_type in meal_plan['days']:
            recipe = self.db.get_recipe_details(recipe_id)
            if recipe:
                meals_by_week[week_num].append({
                    'day': day_num,
                    'name': recipe['name'],
                    'category': recipe['dish_category'],
                    'recipe_id': recipe_id
                })
        
        # Title
        ws['A1'] = f"{meal_plan['name']}"
        ws['A1'].font = Font(bold=True, size=14)
        
        # Week headers (organize by 4-week cycles)
        current_row = 3
        cycle_num = 1
        weeks_in_cycle = 4
        
        for start_week in range(1, meal_plan['num_weeks'] + 1, weeks_in_cycle):
            # Cycle title
            end_week = min(start_week + weeks_in_cycle - 1, meal_plan['num_weeks'])
            cycle_title = f"VIIKKO {start_week}-{end_week}"
            ws[f'A{current_row}'] = cycle_title
            ws[f'A{current_row}'].font = Font(bold=True, size=12)
            current_row += 1
            
            # Draw week columns
            week_col_start = 2
            for week_offset in range(weeks_in_cycle):
                week_num = start_week + week_offset
                if week_num > meal_plan['num_weeks']:
                    break
                
                col = week_col_start + week_offset
                ws.cell(current_row, col).value = f"VIIKKO {week_num}"
                ws.cell(current_row, col).font = Font(bold=True)
                ws.cell(current_row, col).alignment = Alignment(horizontal='center')
            
            current_row += 1
            
            # Draw meal rows (19 per week, all lunch in this version)
            max_meals = max(len(meals_by_week.get(w, [])) for w in range(start_week, end_week + 1))
            max_meals = max(max_meals, 19)  # Ensure at least 19 rows
            
            for meal_num in range(1, max_meals + 1):
                # Meal number label
                ws.cell(current_row, 1).value = f"Meal {meal_num}"
                
                # Add meals for each week in cycle
                for week_offset in range(weeks_in_cycle):
                    week_num = start_week + week_offset
                    if week_num > meal_plan['num_weeks']:
                        continue
                    
                    col = week_col_start + week_offset
                    
                    # Find meal for this week/day combo
                    week_meals = meals_by_week.get(week_num, [])
                    if meal_num <= len(week_meals):
                        meal = week_meals[meal_num - 1]
                        cell = ws.cell(current_row, col)
                        cell.value = meal['name']
                        
                        # Apply category color
                        color = self.CATEGORY_COLORS.get(meal['category'], 'FFFFFF')
                        cell.fill = PatternFill(start_color=color, 
                                              end_color=color, 
                                              fill_type='solid')
                        
                        # Formatting
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        cell.font = Font(size=9)
                
                current_row += 1
            
            # Add spacing between cycles
            current_row += 2
        
        # Add legend
        legend_row = current_row + 1
        ws[f'A{legend_row}'] = "LEGEND:"
        ws[f'A{legend_row}'].font = Font(bold=True)
        
        legend_row += 1
        for category, color in self.CATEGORY_COLORS.items():
            cell = ws.cell(legend_row, 1)
            cell.value = category
            cell.fill = PatternFill(start_color=color, 
                                   end_color=color, 
                                   fill_type='solid')
            legend_row += 1
    
    def _create_ingredient_legend(self, wb, meal_plan):
        """Create ingredient legend sheet showing recipe-ingredient mapping"""
        
        ws = wb.create_sheet("Ingredients")
        
        # Headers
        ws['A1'] = "Recipe"
        ws['B1'] = "Ingredients"
        for col in ['A', 'B']:
            ws[f'{col}1'].font = Font(bold=True)
        
        # Collect all unique recipes
        unique_recipes = set()
        meals_by_week = defaultdict(list)
        for week_num, day_num, recipe_id, meal_type in meal_plan['days']:
            unique_recipes.add(recipe_id)
        
        # List recipes and ingredients
        current_row = 2
        for recipe_id in sorted(unique_recipes):
            recipe = self.db.get_recipe_details(recipe_id)
            if recipe:
                ws.cell(current_row, 1).value = recipe['name']
                
                # Ingredients list
                ingredients_text = ', '.join(
                    f"{ing[0]} ({ing[1]} {ing[2]})" 
                    for ing in recipe['ingredients']
                ) if recipe['ingredients'] else "N/A"
                
                ws.cell(current_row, 2).value = ingredients_text
                ws.cell(current_row, 2).alignment = Alignment(wrap_text=True)
                
                current_row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 60
    
    def _create_recipe_details(self, wb, meal_plan):
        """Create detailed recipe information sheet"""
        
        ws = wb.create_sheet("Recipe Details")
        
        # Headers
        headers = ['Recipe Name', 'Category', 'Ingredients', 'Nutrition', 'Prep Time (min)', 'Difficulty']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(1, col)
            cell.value = header
            cell.font = Font(bold=True)
        
        # Collect all unique recipes
        unique_recipes = set()
        for week_num, day_num, recipe_id, meal_type in meal_plan['days']:
            unique_recipes.add(recipe_id)
        
        # List recipe details
        current_row = 2
        for recipe_id in sorted(unique_recipes):
            recipe = self.db.get_recipe_details(recipe_id)
            if recipe:
                ws.cell(current_row, 1).value = recipe['name']
                ws.cell(current_row, 2).value = recipe['dish_category']
                
                # Ingredients
                ingredients_text = '; '.join(
                    f"{ing[0]}: {ing[1]} {ing[2]}" 
                    for ing in recipe['ingredients']
                ) if recipe['ingredients'] else "N/A"
                ws.cell(current_row, 3).value = ingredients_text
                
                # Nutrition
                if recipe['nutrition']:
                    nutrition_text = ', '.join(
                        f"{k}: {v}" for k, v in recipe['nutrition'].items()
                    )
                    ws.cell(current_row, 4).value = nutrition_text
                
                ws.cell(current_row, 5).value = recipe['prep_time']
                ws.cell(current_row, 6).value = recipe['difficulty']
                
                current_row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 60
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 12


if __name__ == '__main__':
    exporter = MealPlanExporter()
    print("✅ Meal Plan Exporter initialized")

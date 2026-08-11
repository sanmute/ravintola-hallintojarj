"""
Ruokalistasuunnittelija - Meal Plan Generator for Restaurants
Flask web app: recipes -> review -> database -> meal plan -> Excel/Kespro export

Run: python app.py  ->  http://localhost:5001
"""

import os
import sys

# PyInstaller-paketti käynnistetään console=False-tilassa (ei mustaa
# komentoikkunaa) — silloin sys.stdout/sys.stderr (JA sys.__stdout__/
# sys.__stderr__) ovat None heti prosessin käynnistyessä, ja mikä tahansa
# print()-kutsu (näitä on kymmeniä ympäri koodikantaa, mm. koko
# ruokalistageneraattori) kaataa pyynnön: AttributeError: 'NoneType' object
# has no attribute 'write'. Tämä ei koskaan näy kehityskäytössä (python app.py
# ajetaan aina oikeasta konsolista), vain valmiissa .exe:ssä — siksi tämän
# pitää olla ihan ensimmäinen asia joka ajetaan, ENNEN flaskin tai minkään
# muun kirjaston tuontia: monet kirjastot (mm. Werkzeug) kiinnittävät oman
# lokituskahvansa senhetkiseen sys.stderr-arvoon jo tuontihetkellä, joten
# myöhemmin tehty uudelleenohjaus ei enää auttaisi niiden omaa lokitusta.
if sys.stdout is None or sys.stderr is None:
    _log_dir = os.path.join(os.environ.get('LOCALAPPDATA') or os.path.expanduser('~'),
                             'Ruokalistasuunnittelija')
    os.makedirs(_log_dir, exist_ok=True)
    _log_file = open(os.path.join(_log_dir, 'app.log'), 'a', encoding='utf-8', buffering=1)
    sys.stdout = sys.stderr = _log_file
else:
    # Kehityskäytössä (python app.py / KAYNNISTA.bat) konsoli on olemassa,
    # mutta sen merkistö ei suomalaisella/EU-Windowsilla ole aina UTF-8
    # (usein cp1252) — koodikannassa on kymmeniä print()-kutsuja joissa on
    # emoji/erikoismerkkejä (mm. koko ruokalistageneraattori JA
    # meal_plan_modifier.change_meal — yksi eniten käytetyistä ominaisuuksista,
    # "vaihda ateria"), ja ne kaatuvat UnicodeEncodeError-virheeseen ei-UTF8
    # konsolissa. Pakota UTF-8 ja korvaa tulostuskelvottomat merkit sen sijaan
    # että pyyntö kaatuu.
    for _stream in (sys.stdout, sys.stderr):
        try:
            _stream.reconfigure(encoding='utf-8', errors='replace')
        except (AttributeError, ValueError):
            pass

import json
import sqlite3
import threading
import time
from datetime import datetime, timedelta
from collections import defaultdict

from flask import Flask, render_template, request, jsonify, send_file

from meal_plan_db import MealPlanDB
from meal_plan_generator import MealPlanGenerator
from meal_plan_exporter import MealPlanExporter
from meal_plan_modifier import MealModifier

from backup import init_backup
from translator import translate_en_to_fi

import shutil
import pytesseract
# Älä kovakoodaa kehittäjän omaa polkua — se ei ole olemassa keittiön koneella
# eri käyttäjätunnuksen alla, jolloin OCR-tuonti olisi rikki joka asennuksessa
# paitsi tällä yhdellä kehityskoneella. Etsi sen sijaan PATH:sta ja yleisimmistä
# asennuspaikoista; jos mitään ei löydy, jätetään pytesseractin oletus voimaan
# ja virhe näkyy vasta jos/kun OCR-ominaisuutta oikeasti käytetään.
_tesseract_candidates = [
    shutil.which('tesseract'),
    r'C:\Program Files\Tesseract-OCR\tesseract.exe',
    r'C:\Program Files (x86)\Tesseract-OCR\tesseract.exe',
    os.path.join(os.environ.get('LOCALAPPDATA', ''), 'Tesseract-OCR', 'tesseract.exe'),
]
for _candidate in _tesseract_candidates:
    if _candidate and os.path.exists(_candidate):
        pytesseract.pytesseract.tesseract_cmd = _candidate
        break

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 64 * 1024 * 1024  # 64 MB

_APP_START_TIME = time.time()

def _user_data_dir():
    """Pysyvä tallennuskansio käyttäjäkohtaiselle datalle.

    Ei koskaan sovelluksen asennus-/build-kansiossa: PyInstaller-paketin
    _internal-kansio korvataan kokonaan jokaisen uudelleenkäännöksen tai
    -asennuksen yhteydessä, jolloin tietokanta katoaisi joka käynnistyksellä.
    """
    base = os.environ.get('LOCALAPPDATA') or os.path.expanduser('~')
    data_dir = os.path.join(base, 'Ruokalistasuunnittelija')
    os.makedirs(data_dir, exist_ok=True)
    return data_dir


_DATA_DIR = _user_data_dir()
DB_PATH = os.path.join(_DATA_DIR, 'meal_plans.db')
OUTPUT_DIR = os.path.join(_DATA_DIR, 'exports')
os.makedirs(OUTPUT_DIR, exist_ok=True)


def _downloads_dir():
    """Käyttäjän oikea Windows-Lataukset-kansio (~\\Downloads). Palvelin ja
    työpöytäsovelluksen ikkuna ajavat samalla koneella, joten tiedostot voi
    kirjoittaa suoraan sinne sen sijaan että luotettaisiin selaimen
    lataustoimintoon — pywebviewin sisäinen ikkuna ei tue sitä ollenkaan."""
    path = os.path.join(os.path.expanduser('~'), 'Downloads')
    os.makedirs(path, exist_ok=True)
    return path


def _save_to_downloads(src_path, download_name):
    """Kopioi valmiin vientitiedoston Lataukset-kansioon nimellä download_name
    (lisää ' (2)' tms. jos samanniminen tiedosto on jo olemassa, ettei
    ylikirjoiteta käyttäjän aiempaa vientiä huomaamatta). Palauttaa lopullisen
    polun."""
    import shutil as _shutil
    base, ext = os.path.splitext(download_name)
    dest_dir = _downloads_dir()
    dest = os.path.join(dest_dir, download_name)
    n = 2
    while os.path.exists(dest):
        dest = os.path.join(dest_dir, f'{base} ({n}){ext}')
        n += 1
    _shutil.copy2(src_path, dest)
    return dest


# Siirrä vanha tietokanta automaattisesti, jos sellainen löytyy sovelluksen
# vieressä (aiempi tallennuspaikka) eikä uudessa sijainnissa ole vielä dataa.
if not os.path.exists(DB_PATH):
    _legacy_db = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'meal_plans.db')
    if os.path.exists(_legacy_db) and os.path.abspath(_legacy_db) != os.path.abspath(DB_PATH):
        import shutil as _shutil
        _shutil.copy2(_legacy_db, DB_PATH)

MealPlanDB(DB_PATH)  # ensure core tables exist before anything else touches DB_PATH

with sqlite3.connect(DB_PATH) as _startup_conn:
    _cols = [r[1] for r in _startup_conn.execute('PRAGMA table_info(recipes)').fetchall()]
    if 'manual_only' not in _cols:
        _startup_conn.execute('ALTER TABLE recipes ADD COLUMN manual_only INTEGER DEFAULT 0')
    if 'servings' not in _cols:
        _startup_conn.execute('ALTER TABLE recipes ADD COLUMN servings INTEGER')
    if 'recipe_type' not in _cols:
        # Korvaa generaattorin aiemman nimipohjaisen keitto/salaatti-tunnistuksen
        # (SOUP_KEYWORD/SALAD_KEYWORD-merkkijonohaku nimestä) eksplisiittisellä
        # sarakkeella. Olemassa olevat reseptit luokitellaan kertaalleen samalla
        # logiikalla jota generaattori käytti, jotta käyttäytyminen ei muutu —
        # tästä eteenpäin luokitus on pysyvä ja muokattavissa käyttöliittymässä.
        _startup_conn.execute("ALTER TABLE recipes ADD COLUMN recipe_type TEXT DEFAULT 'pääruoka'")
        _startup_conn.execute(
            "UPDATE recipes SET recipe_type='salaatti' WHERE lower(name_fi) LIKE '%salaa%'")
        _startup_conn.execute(
            "UPDATE recipes SET recipe_type='keitto' "
            "WHERE lower(name_fi) LIKE '%keitto%' AND lower(name_fi) NOT LIKE '%salaa%'")
    if 'instructions' not in _cols:
        # Valmistusohjeet napattiin talteen jo aiemmin tuonnin/tarkistuksen
        # aikana (instructions_raw skräpatuissa/OCR:atuissa/Oma resepti
        # -reseptelyissä), mutta niitä ei koskaan tallennettu itse
        # recipes-tauluun — ne hävisivät hyväksynnän yhteydessä. Tästä
        # eteenpäin ne tallennetaan pysyvästi.
        _startup_conn.execute("ALTER TABLE recipes ADD COLUMN instructions TEXT")

    _mp_cols = [r[1] for r in _startup_conn.execute('PRAGMA table_info(meal_plans)').fetchall()]
    if 'facility' not in _mp_cols:
        # Sama reseptipohja kaikille toimipisteille (kesti/hoiva/kymenkartano),
        # mutta jokaiselle oma erikseen generoitu ruokalista — kesti/hoiva
        # käyttävät samoja reseptejä ma-pe, hoiva lisäksi la-su; kymenkartano
        # saa oman erillisen suunnitelman samasta reseptipoolista. Olemassa
        # olevat suunnitelmat ovat kaikki 'kesti' (ainoa toimipiste ennen tätä).
        _startup_conn.execute("ALTER TABLE meal_plans ADD COLUMN facility TEXT DEFAULT 'kesti'")

    if 'mirrors_plan_id' not in _mp_cols:
        # Hoiva EI ole enää itsenäisesti generoitu — sen ma-pe on aina
        # tasan sama kuin linkitetyn kesti-suunnitelman (vahvistettu
        # käyttäjältä), ja vain la-su on hoiva-suunnitelman omaa dataa.
        # Kymenkartano pysyy täysin itsenäisenä (fyysisesti eri paikka).
        _startup_conn.execute("ALTER TABLE meal_plans ADD COLUMN mirrors_plan_id INTEGER")

    _mpd_cols = [r[1] for r in _startup_conn.execute('PRAGMA table_info(meal_plan_days)').fetchall()]
    if 'dessert_text' not in _mpd_cols:
        # Vapaamuotoinen jälkiruokateksti perjantaille (day_of_week=4, aina
        # perjantai riippumatta toimipisteestä koska maanantai=0). Tallennetaan
        # päivän 'lounas'-riville, koska jokaisella päivällä on aina tasan yksi.
        _startup_conn.execute("ALTER TABLE meal_plan_days ADD COLUMN dessert_text TEXT")

    # "Uudet reseptit" -jako: kaikki reseptit jotka olivat jo tietokannassa
    # ennen tätä ominaisuutta ovat pysyvästi "legacy" — kaikki tästä eteenpäin
    # lisätyt (mistä tahansa lähteestä: käsin, PoweResta, OCR, verkkokaappaus,
    # myös vielä kirjoittamattomat tuontitavat) ovat pysyvästi "uusia".
    # Rajapyykki on yksinkertaisesti "suurin resepti-id tällä hetkellä" —
    # jokainen tuontireitti kasvattaa id:tä AUTOINCREMENTilla, joten mikään
    # koodimuutos ei ole tarpeen yksittäisissä tuontirouteissa.
    _startup_conn.execute('''CREATE TABLE IF NOT EXISTS app_settings (
        key TEXT PRIMARY KEY, value TEXT
    )''')
    _existing_cutover = _startup_conn.execute(
        "SELECT value FROM app_settings WHERE key = 'recipe_overhaul_cutover_id'").fetchone()
    if _existing_cutover is None:
        _max_id = _startup_conn.execute('SELECT COALESCE(MAX(id), 0) FROM recipes').fetchone()[0]
        _startup_conn.execute(
            "INSERT INTO app_settings (key, value) VALUES ('recipe_overhaul_cutover_id', ?)",
            (str(_max_id),))


def _recipe_overhaul_cutover_id(conn):
    row = conn.execute(
        "SELECT value FROM app_settings WHERE key = 'recipe_overhaul_cutover_id'").fetchone()
    return int(row[0]) if row else 0

init_backup(app, DB_PATH)

from order_catalog import init_order_catalog 
init_order_catalog(app, DB_PATH)

SEASONS = {'talvi': 'Talvi', 'kevät': 'Kevät', 'kesä': 'Kesä', 'syksy': 'Syksy', 'kaikki': 'Ympärivuotinen'}
CATEGORIES = {'kala': 'Kala', 'kana': 'Kana', 'naudanliha': 'Liha', 'kasvis': 'Kasvis'}
CATEGORY_COLORS = {'kala': '#4A90D9', 'kana': '#F5D547', 'naudanliha': '#D94A4A', 'kasvis': '#5CB85C'}
RECIPE_TYPES = ('pääruoka', 'keitto', 'salaatti', 'leivonta', 'jälkiruoka',
                'kastike', 'kasvislisäke', 'energialisäke')
# Nämä kolme koskevat VAIN Kestiä (ei Hoivaa, ei Kymenkartanoa) — vapaaehtoisia
# lisäyksiä joita generaattori ei koskaan valitse itse, ks. ACCOMPANIMENT_CONFIG.
KESTI_ONLY_RECIPE_TYPES = ('kastike', 'kasvislisäke', 'energialisäke')

# ---- Tietoa-sivun kiinteät tiedot — täytä oikeat yhteystiedot ennen käyttöönottoa ----
APP_VERSION = '1.0.0'
APP_RELEASE_DATE = '16.7.2026'
DEVELOPER_NAME = 'Santeri Mutanen'
DEVELOPER_EMAIL = 'semutanen@gmail.com'
DEVELOPER_PHONE = '+358442012530'
ORGANIZATION = 'Jyränkölän Setlementti Ry'
IT_CONTACT = 'jussi.varjo@jyrankola.fi'
ADMIN_NAME = 'Teijo Lehikoinen'
ADMIN_PHONE = '+358447972471'


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


class _BadIntArg(Exception):
    def __init__(self, message):
        self.message = message


def _int_arg(args, key, default=None):
    """Parse an integer query/JSON argument, raising _BadIntArg (caught by the
    caller into a clean 400 response) instead of letting a typo'd or missing
    value crash the request with an unhandled ValueError/TypeError — a raw
    500 page is especially bad here since several callers reach these routes
    via a full-page navigation (window.location), which would blank out the
    entire app for a non-technical user over one bad input."""
    raw = args.get(key, default)
    if raw is None:
        raise _BadIntArg(f"Puuttuva tai virheellinen parametri '{key}'")
    try:
        return int(raw)
    except (TypeError, ValueError):
        raise _BadIntArg(f"Puuttuva tai virheellinen parametri '{key}'")


def _optional_int_arg(args, key):
    """Like _int_arg, but a genuinely absent value returns None instead of
    raising — for params that are optional by design (e.g. 'servings' on
    exports that work fine unscaled). Still raises _BadIntArg for a
    present-but-unparseable value."""
    raw = args.get(key)
    if raw is None or raw == '':
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        raise _BadIntArg(f"Virheellinen parametri '{key}'")


def _code_last_updated():
    """Uusin muokkausajankohta projektin .py-tiedostoista — käytetään
    'Päivitetty'-tietona Tietoa-sivulla."""
    root = os.path.dirname(os.path.abspath(__file__))
    py_files = [f for f in os.listdir(root) if f.endswith('.py')]
    newest = max((os.path.getmtime(os.path.join(root, f)) for f in py_files), default=None)
    if newest is None:
        return APP_RELEASE_DATE
    return datetime.fromtimestamp(newest).strftime('%d.%m.%Y')


@app.route('/')
def index():
    return render_template('index.html', seasons=SEASONS, categories=CATEGORIES,
                           category_colors=CATEGORY_COLORS,
                           app_version=APP_VERSION, app_release_date=APP_RELEASE_DATE,
                           app_updated=_code_last_updated(),
                           developer_name=DEVELOPER_NAME, developer_email=DEVELOPER_EMAIL,
                           developer_phone=DEVELOPER_PHONE, organization=ORGANIZATION,
                           it_contact=IT_CONTACT, admin_name=ADMIN_NAME, admin_phone=ADMIN_PHONE)


# ============================================================
# RECIPES: list / add / edit / delete / review queue
# ============================================================

@app.route('/api/recipes')
def list_recipes():
    season = request.args.get('season', '')
    conn = get_db()
    q = ('SELECT id, name_fi, season, meal_type, dish_category, recipe_type, notes, '
         'manual_only, instructions, servings FROM recipes')
    params = []
    if season:
        q += ' WHERE season = ?'
        params.append(season)
    q += ' ORDER BY dish_category, name_fi'
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/recipes', methods=['POST'])
def add_recipe():
    data = request.json or {}
    name = (data.get('name_fi') or '').strip()
    season = data.get('season')
    category = data.get('dish_category')
    if not name or season not in SEASONS or category not in CATEGORIES:
        return jsonify({'error': 'Nimi, kausi ja kategoria vaaditaan'}), 400
    recipe_type = data.get('recipe_type') if data.get('recipe_type') in RECIPE_TYPES else 'pääruoka'
    servings = None
    if data.get('servings') not in (None, ''):
        try:
            servings = int(data['servings'])
            if not (1 <= servings <= 500):
                servings = None
        except (TypeError, ValueError):
            servings = None
    db = MealPlanDB(DB_PATH)
    recipe_id = db.add_recipe(
        name_fi=name, season=season,
        meal_type=data.get('meal_type', 'lounas'),
        dish_category=category,
        recipe_type=recipe_type,
        notes=data.get('notes', ''),
        instructions=data.get('instructions', ''),
        servings=servings
    )
    if recipe_id is None:
        return jsonify({'error': f"Resepti '{name}' on jo olemassa"}), 409
    conn = get_db()
    _ensure_recipe_source_column(conn)
    _set_recipe_source(conn, recipe_id, 'manual')
    conn.commit()
    conn.close()
    return jsonify({'id': recipe_id, 'message': f"Resepti '{name}' lisätty"})


@app.route('/api/recipes/<int:recipe_id>', methods=['PUT'])
def update_recipe(recipe_id):
    data = request.json or {}
    conn = get_db()
    fields, params = [], []
    for col in ('name_fi', 'season', 'meal_type', 'dish_category', 'notes', 'instructions'):
        if col in data:
            fields.append(f'{col} = ?')
            params.append(data[col])
    if 'recipe_type' in data:
        if data['recipe_type'] not in RECIPE_TYPES:
            conn.close()
            return jsonify({'error': 'Virheellinen tyyppi'}), 400
        fields.append('recipe_type = ?')
        params.append(data['recipe_type'])
    if 'manual_only' in data:
        fields.append('manual_only = ?')
        params.append(1 if data['manual_only'] else 0)
    if 'servings' in data:
        if data['servings'] in (None, ''):
            fields.append('servings = ?')
            params.append(None)
        else:
            try:
                servings = int(data['servings'])
            except (TypeError, ValueError):
                conn.close()
                return jsonify({'error': 'Annosmäärän pitää olla numero'}), 400
            if not (1 <= servings <= 500):
                conn.close()
                return jsonify({'error': 'Annosmäärän pitää olla 1-500'}), 400
            fields.append('servings = ?')
            params.append(servings)
    if not fields:
        conn.close()
        return jsonify({'error': 'Ei muutoksia'}), 400
    params.append(recipe_id)
    try:
        # Legacy/new is a permanent id-based cutover (see
        # _recipe_overhaul_cutover_id) — editing a legacy recipe does not
        # move it into "new".
        conn.execute(f'UPDATE recipes SET {", ".join(fields)} WHERE id = ?', params)
        conn.commit()
    except sqlite3.IntegrityError:
        conn.close()
        return jsonify({'error': 'Samanniminen resepti on jo olemassa'}), 409
    conn.close()
    return jsonify({'message': 'Resepti päivitetty'})


@app.route('/api/recipes/<int:recipe_id>', methods=['DELETE'])
def delete_recipe(recipe_id):
    """Poistaminen menee aina läpi, vaikka resepti olisi käytössä jollain
    ruokalistalla — poisto ei odota suunnitelman muokkausta. Käytössä olleet
    ruokalistapaikat poistetaan mukana, ja käyttäjää pyydetään uusimaan
    kyseiset suunnitelmat (Ruokalistat-välilehti)."""
    conn = get_db()
    affected_plans = conn.execute(
        '''SELECT DISTINCT mp.name FROM meal_plan_days d
           JOIN meal_plans mp ON mp.id = d.meal_plan_id
           WHERE d.recipe_id = ?''', (recipe_id,)).fetchall()
    conn.execute('DELETE FROM meal_plan_days WHERE recipe_id = ?', (recipe_id,))
    conn.execute('DELETE FROM recipe_ingredients WHERE recipe_id = ?', (recipe_id,))
    conn.execute('DELETE FROM ingredient_changes WHERE recipe_id = ?', (recipe_id,))
    conn.execute('DELETE FROM recipes WHERE id = ?', (recipe_id,))
    conn.commit()
    conn.close()

    message = 'Resepti poistettu.'
    if affected_plans:
        names = ', '.join(r['name'] for r in affected_plans)
        message += f' Se oli käytössä ruokalistoilla: {names}. Uusi kyseiset suunnitelmat Ruokalistat-välilehdeltä.'
    return jsonify({'message': message, 'affected_plans': [r['name'] for r in affected_plans]})


# ============================================================
# RECIPE AUDIT: tunnista jälkiruoat/juomat/välipalat/pelkät lisäkkeet,
# jotka eivät kuulu pääruoka-, keitto- tai salaattikiertoon
# ============================================================

@app.route('/api/admin/audit-recipes')
def audit_recipes():
    """Listaa reseptit, jotka eivät todennäköisesti kuulu automaattiseen
    generointiin (jälkiruoat, juomat, välipalat, pelkät lisäkkeet), luokan
    mukaan ryhmiteltynä. Keitot ja salaatit jätetään kokonaan pois — ne
    ovat jo omassa, toimivaksi todetussa kierrossaan."""
    from recipe_classifier import classify_recipe

    conn = get_db()
    recipes = conn.execute(
        'SELECT id, name_fi, dish_category FROM recipes ORDER BY name_fi').fetchall()
    conn.close()

    classified = {'dessert': [], 'drink': [], 'snack': [], 'side_only': [], 'unknown': []}
    for r in recipes:
        category = classify_recipe(r['name_fi'])
        if category is None:
            continue  # soup/salad — already in active rotation, not audited
        classified[category].append({
            'id': r['id'], 'name_fi': r['name_fi'], 'dish_category': r['dish_category']
        })

    summary = {
        'desserts': len(classified['dessert']),
        'drinks': len(classified['drink']),
        'snacks': len(classified['snack']),
        'sides': len(classified['side_only']),
        'unknown': len(classified['unknown']),
        'total': len(recipes),
    }
    return jsonify({'summary': summary, 'recipes': classified})


@app.route('/api/admin/delete-recipes', methods=['POST'])
def delete_recipes_bulk():
    """Poista merkityt reseptit. Poisto menee aina läpi, vaikka resepti
    olisi käytössä jollain ruokalistalla — käytössä olleet ruokalistapaikat
    poistetaan mukana, eikä käyttäjän tarvitse ensin käsin vaihtaa niitä.
    Vaikutetut suunnitelmat listataan, jotta ne osataan uusia."""
    data = request.json or {}
    recipe_ids = [int(rid) for rid in data.get('recipe_ids', [])]
    if not recipe_ids:
        return jsonify({'error': 'Ei valittuja reseptejä'}), 400

    conn = get_db()
    deleted = []
    affected_plans = set()
    for recipe_id in recipe_ids:
        row = conn.execute('SELECT name_fi FROM recipes WHERE id = ?', (recipe_id,)).fetchone()
        if not row:
            continue
        plans = conn.execute(
            '''SELECT DISTINCT mp.name FROM meal_plan_days d
               JOIN meal_plans mp ON mp.id = d.meal_plan_id
               WHERE d.recipe_id = ?''', (recipe_id,)).fetchall()
        affected_plans.update(r['name'] for r in plans)
        conn.execute('DELETE FROM meal_plan_days WHERE recipe_id = ?', (recipe_id,))
        conn.execute('DELETE FROM recipe_ingredients WHERE recipe_id = ?', (recipe_id,))
        conn.execute('DELETE FROM ingredient_changes WHERE recipe_id = ?', (recipe_id,))
        conn.execute('DELETE FROM recipes WHERE id = ?', (recipe_id,))
        deleted.append(recipe_id)
    conn.commit()
    conn.close()

    message = f'Poistettu {len(deleted)} reseptiä.'
    if affected_plans:
        message += (f' Vaikutti ruokalistoihin: {", ".join(sorted(affected_plans))}. '
                    f'Uusi kyseiset suunnitelmat Ruokalistat-välilehdeltä.')
    return jsonify({'success': True, 'deleted_count': len(deleted),
                    'affected_plans': sorted(affected_plans), 'message': message})


@app.route('/api/recipes/list-by-source')
def list_recipes_by_source():
    """Jaa reseptit 'vanhoihin' (olivat tietokannassa jo ennen tätä
    ominaisuutta) ja 'uusiin' (lisätty sen jälkeen, mistä tahansa lähteestä —
    käsin, PoweResta, OCR, verkkokaappaus). Jako on pysyvä id-rajapyykki,
    ei riipu tuontikanavasta eikä muutu myöhemmin tehdyistä muokkauksista."""
    conn = get_db()
    _ensure_recipe_source_column(conn)
    cutover_id = _recipe_overhaul_cutover_id(conn)
    new_rows = conn.execute(
        '''SELECT id, name_fi, season, dish_category, source, created_at FROM recipes
           WHERE id > ? ORDER BY created_at DESC''', (cutover_id,)).fetchall()
    legacy_rows = conn.execute(
        '''SELECT id, name_fi, season, dish_category, source, created_at FROM recipes
           WHERE id <= ? ORDER BY name_fi''', (cutover_id,)).fetchall()
    conn.close()

    return jsonify({
        'legacy': [dict(r) for r in legacy_rows],
        'new': [dict(r) for r in new_rows],
        'legacy_count': len(legacy_rows),
        'new_count': len(new_rows),
    })


@app.route('/api/recipes/export-poweresta', methods=['POST'])
def export_poweresta():
    """Vie valitut reseptit yhdeksi PoweResta-muotoiseksi Excel-tiedostoksi
    (yksi resepti per välilehti). Raaka-aineet haetaan samasta paikasta kuin
    reseptin muokkausnäkymässä (recipe_ingredients-liitostaulu) — reseptit.
    ingredients-JSON-sarake ei koskaan täyty millään tuontireitillä tässä
    sovelluksessa, joten sen lukeminen tuottaisi aina tyhjät reseptit."""
    from menu_pdf_generator import extract_menu_codes
    from poweresta_exporter import PoweRestaExporter

    data = request.json or {}
    recipe_ids = [int(rid) for rid in data.get('recipe_ids', [])]
    if not recipe_ids:
        return jsonify({'error': 'Ei valittuja reseptejä'}), 400

    conn = get_db()
    exporter = PoweRestaExporter()
    exported = 0
    for recipe_id in recipe_ids:
        recipe = conn.execute(
            'SELECT name_fi, servings, notes FROM recipes WHERE id = ?', (recipe_id,)).fetchone()
        if not recipe:
            continue
        ing_rows = conn.execute(
            '''SELECT i.name_fi AS name, ri.quantity, ri.unit
               FROM recipe_ingredients ri JOIN ingredients i ON i.id = ri.ingredient_id
               WHERE ri.recipe_id = ? ORDER BY i.name_fi''', (recipe_id,)).fetchall()
        exporter.add_recipe(
            recipe['name_fi'],
            [dict(r) for r in ing_rows],
            servings=recipe['servings'],
            diets=extract_menu_codes(recipe['notes']).strip('() ') or None,
        )
        exported += 1
    conn.close()

    if not exported:
        return jsonify({'error': 'Valituista resepteistä yhtäkään ei löytynyt'}), 404

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'poweresta_vienti_{timestamp}.xlsx'
    return send_file(exporter.get_bytes(), as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================================
# RECIPE SOURCE TRACKING: käytetään raaka-aineiden tarkistusraportissa
# ============================================================

USER_CONTRIBUTED_MARK = '[käyttäjän lisäämä]'


def _ensure_recipe_source_column(conn):
    """Lisää recipes.source ja recipes.ingredients_audit_ok -sarakkeet jos
    puuttuvat, ja paikkaa olemassa olevat rivit parhaalla mahdollisella
    tiedolla. Reseptin alkuperäistä tuontikanavaa (PoweResta vs. OCR vs.
    verkkokaappaus) EI tallennettu ennen tätä muutosta — review_approve()
    hylkäsi sen tiedon ennen tätä. Siksi jo olemassa olevat reseptit,
    joiden alkuperää ei voida luotettavasti päätellä jälkikäteen, merkitään
    'tuntematon (ennen lähdeseurantaa)' sen sijaan että arvattaisiin."""
    cols = [r[1] for r in conn.execute('PRAGMA table_info(recipes)').fetchall()]
    if 'source' not in cols:
        conn.execute('ALTER TABLE recipes ADD COLUMN source TEXT')
    if 'ingredients_audit_ok' not in cols:
        conn.execute('ALTER TABLE recipes ADD COLUMN ingredients_audit_ok INTEGER DEFAULT 0')
    conn.execute(
        "UPDATE recipes SET source='user_contributed' WHERE source IS NULL AND notes LIKE ?",
        (f'%{USER_CONTRIBUTED_MARK}%',))
    conn.execute(
        "UPDATE recipes SET source='legacy_tuntematon' WHERE source IS NULL")


def _set_recipe_source(conn, recipe_id, source):
    if recipe_id is not None:
        conn.execute('UPDATE recipes SET source=? WHERE id=?', (source, recipe_id))



def _classify_import_source(source_site, source_url):
    """Päättele tuontikanava tarkistusjonon 'source_site'/'source_url'-kentistä
    (ks. import_poweresta, ocr_menu_import.menu_to_queue_items, import_paste,
    recipe_scraper_v2.scrape_recipe)."""
    s = (source_site or '').lower()
    if s.startswith('poweresta'):
        return 'poweresta'
    if s.startswith('ocr'):
        return 'ocr'
    if 'käsin' in s:
        return 'manual_paste'
    if source_url and source_url.startswith('http'):
        return 'scraped'
    return 'tuntematon'


# ============================================================
# RECIPE INGREDIENTS: yhteismuokkaus muutoshistorialla
# ============================================================

def _ensure_ingredient_editor_tables(conn):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS ingredients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name_fi TEXT UNIQUE NOT NULL);
        CREATE TABLE IF NOT EXISTS recipe_ingredients (
            recipe_id INTEGER NOT NULL,
            ingredient_id INTEGER NOT NULL,
            quantity REAL, unit TEXT,
            PRIMARY KEY (recipe_id, ingredient_id));
        CREATE TABLE IF NOT EXISTS ingredient_changes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            recipe_id INTEGER NOT NULL,
            changed_by_user_id INTEGER,
            changed_by_username TEXT,
            change_type TEXT NOT NULL,
            ingredient_name TEXT NOT NULL,
            old_quantity REAL, new_quantity REAL,
            old_unit TEXT, new_unit TEXT,
            timestamp TEXT DEFAULT (datetime('now')),
            notes TEXT
        );
    """)
    cols = [r[1] for r in conn.execute('PRAGMA table_info(recipe_ingredients)').fetchall()]
    if 'ingredient_instruction' not in cols:
        conn.execute('ALTER TABLE recipe_ingredients ADD COLUMN ingredient_instruction TEXT')
    if 'sort_order' not in cols:
        # Mahdollistaa käsin järjestämisen sen sijaan että raaka-aineet
        # näytettäisiin aina aakkosjärjestyksessä. Olemassa olevat rivit
        # varustetaan nykyisellä aakkosjärjestyksellään, jotta mikään ei
        # hyppää näkyvästi paikaltaan tämän muutoksen myötä.
        conn.execute('ALTER TABLE recipe_ingredients ADD COLUMN sort_order INTEGER DEFAULT 0')
        _backfill_rows = conn.execute(
            '''SELECT ri.recipe_id, ri.ingredient_id
               FROM recipe_ingredients ri JOIN ingredients i ON i.id = ri.ingredient_id
               ORDER BY ri.recipe_id, i.name_fi''').fetchall()
        _counters = {}
        for recipe_id, ingredient_id in _backfill_rows:
            _counters[recipe_id] = _counters.get(recipe_id, 0) + 1
            conn.execute(
                'UPDATE recipe_ingredients SET sort_order = ? WHERE recipe_id = ? AND ingredient_id = ?',
                (_counters[recipe_id], recipe_id, ingredient_id))


def _log_ingredient_change(conn, recipe_id, change_type, ingredient_name,
                           old_quantity=None, new_quantity=None, old_unit=None, new_unit=None):
    conn.execute(
        '''INSERT INTO ingredient_changes
           (recipe_id, changed_by_user_id, changed_by_username, change_type,
            ingredient_name, old_quantity, new_quantity, old_unit, new_unit)
           VALUES (?,?,?,?,?,?,?,?,?)''',
        (recipe_id, None, 'admin', change_type,
         ingredient_name, old_quantity, new_quantity, old_unit, new_unit))


def _get_or_create_ingredient_id(conn, name):
    conn.execute('INSERT OR IGNORE INTO ingredients (name_fi) VALUES (?)', (name,))
    return conn.execute('SELECT id FROM ingredients WHERE name_fi=?', (name,)).fetchone()[0]


def _describe_ingredient_change(row):
    if row['change_type'] == 'added':
        return f"Lisätty: {row['ingredient_name']} {row['new_quantity'] or ''} {row['new_unit'] or ''}".strip()
    if row['change_type'] == 'removed':
        return f"Poistettu: {row['ingredient_name']}"
    if row['change_type'] == 'audit_checked':
        notes = row['notes'] if 'notes' in row.keys() else None
        return f"Tarkistettu: {notes}" if notes else 'Tarkistettu'
    old = f"{row['old_quantity'] if row['old_quantity'] is not None else '-'} {row['old_unit'] or ''}".strip()
    new = f"{row['new_quantity'] if row['new_quantity'] is not None else '-'} {row['new_unit'] or ''}".strip()
    return f"{row['ingredient_name']}: {old} → {new}"


@app.route('/api/recipes/<int:recipe_id>/details')
def get_recipe_details(recipe_id):
    conn = get_db()
    _ensure_ingredient_editor_tables(conn)
    recipe = conn.execute('SELECT * FROM recipes WHERE id=?', (recipe_id,)).fetchone()
    if not recipe:
        conn.close()
        return jsonify({'error': 'Reseptiä ei löydy'}), 404

    ingredients = conn.execute(
        '''SELECT ri.ingredient_id, i.name_fi AS ingredient_name, ri.quantity, ri.unit,
                  ri.ingredient_instruction, ri.sort_order
           FROM recipe_ingredients ri JOIN ingredients i ON i.id = ri.ingredient_id
           WHERE ri.recipe_id=? ORDER BY ri.sort_order, i.name_fi''', (recipe_id,)).fetchall()

    changes = conn.execute(
        '''SELECT timestamp, changed_by_username, change_type, ingredient_name,
                  old_quantity, new_quantity, old_unit, new_unit, notes
           FROM ingredient_changes WHERE recipe_id=?
           ORDER BY timestamp DESC, id DESC LIMIT 20''', (recipe_id,)).fetchall()
    conn.close()

    recipe_dict = dict(recipe)
    recipe_dict['category'] = recipe_dict.get('dish_category')  # yhdenmukaisuuden vuoksi UI:lle

    return jsonify({
        'recipe': recipe_dict,
        'ingredients': [dict(r) for r in ingredients],
        'change_history': [
            {'timestamp': c['timestamp'], 'changed_by': c['changed_by_username'] or '—',
             'change_type': c['change_type'], 'description': _describe_ingredient_change(c)}
            for c in changes
        ],
    })


@app.route('/api/recipes/<int:recipe_id>/print-pdf')
def print_recipe_pdf(recipe_id):
    """Tulosta yksittäinen resepti PDF:nä keittiökäyttöön. Kuten viikkomenun
    vienti — työpöytäversion sisäinen ikkuna (pywebview/WebView2) ei tue
    selaimen lataustoimintoa, joten tiedosto kirjoitetaan suoraan käyttäjän
    oikeaan Lataukset-kansioon palvelimella (ks. _save_to_downloads)."""
    conn = get_db()
    recipe = conn.execute(
        'SELECT name_fi, recipe_type, season, servings, instructions FROM recipes WHERE id = ?',
        (recipe_id,)).fetchone()
    if not recipe:
        conn.close()
        return jsonify({'error': 'Reseptiä ei löydy'}), 404

    ingredients = conn.execute(
        '''SELECT i.name_fi, ri.quantity, ri.unit, ri.ingredient_instruction
           FROM recipe_ingredients ri JOIN ingredients i ON i.id = ri.ingredient_id
           WHERE ri.recipe_id = ? ORDER BY ri.sort_order, i.name_fi''', (recipe_id,)).fetchall()
    conn.close()

    ingredient_dicts = []
    for ing in ingredients:
        name = ing['name_fi']
        if ing['ingredient_instruction']:
            name = f"{name} ({ing['ingredient_instruction']})"
        ingredient_dicts.append({'name_fi': name, 'quantity': ing['quantity'], 'unit': ing['unit']})

    from menu_pdf_generator import build_single_recipe_pdf
    safe_name = ''.join(c for c in recipe['name_fi'] if c.isalnum() or c in ' -_').strip() or f'resepti_{recipe_id}'
    out = os.path.join(OUTPUT_DIR, f'resepti_{recipe_id}.pdf')
    build_single_recipe_pdf(dict(recipe), ingredient_dicts, out)
    download_name = f'{safe_name}.pdf'
    saved_path = _save_to_downloads(out, download_name)

    return jsonify({
        'message': f'Resepti tallennettu Lataukset-kansioon: {download_name}',
        'filename': download_name,
        'saved_path': saved_path,
    })


@app.route('/api/recipes/<int:recipe_id>/ingredient-changes')
def get_recipe_ingredient_changes(recipe_id):
    conn = get_db()
    _ensure_ingredient_editor_tables(conn)
    changes = conn.execute(
        '''SELECT timestamp, changed_by_username, change_type, ingredient_name,
                  old_quantity, new_quantity, old_unit, new_unit
           FROM ingredient_changes WHERE recipe_id=?
           ORDER BY timestamp DESC, id DESC LIMIT 50''', (recipe_id,)).fetchall()
    conn.close()
    return jsonify({'changes': [dict(c) for c in changes]})


@app.route('/api/recipes/<int:recipe_id>/ingredients', methods=['POST'])
def add_recipe_ingredient(recipe_id):
    d = request.json or {}
    name = (d.get('ingredient_name') or '').strip()
    if not name:
        return jsonify({'error': 'Raaka-aineen nimi vaaditaan'}), 400
    try:
        quantity = round(float(d['quantity']), 3) if d.get('quantity') not in (None, '') else None
    except (TypeError, ValueError):
        return jsonify({'error': 'Virheellinen määrä'}), 400
    unit = (d.get('unit') or '').strip() or None
    instruction = (d.get('ingredient_instruction') or '').strip() or None

    conn = get_db()
    _ensure_ingredient_editor_tables(conn)
    if not conn.execute('SELECT 1 FROM recipes WHERE id=?', (recipe_id,)).fetchone():
        conn.close()
        return jsonify({'error': 'Reseptiä ei löydy'}), 404

    ing_id = _get_or_create_ingredient_id(conn, name)
    existing = conn.execute(
        'SELECT 1 FROM recipe_ingredients WHERE recipe_id=? AND ingredient_id=?',
        (recipe_id, ing_id)).fetchone()
    if existing:
        conn.close()
        return jsonify({'error': f"'{name}' on jo tämän reseptin raaka-aineissa — muokkaa "
                                 f"olemassa olevaa riviä sen sijaan"}), 409

    next_order = conn.execute(
        'SELECT COALESCE(MAX(sort_order), 0) + 1 FROM recipe_ingredients WHERE recipe_id=?',
        (recipe_id,)).fetchone()[0]
    conn.execute(
        '''INSERT INTO recipe_ingredients
           (recipe_id, ingredient_id, quantity, unit, ingredient_instruction, sort_order)
           VALUES (?,?,?,?,?,?)''',
        (recipe_id, ing_id, quantity, unit, instruction, next_order))
    _log_ingredient_change(conn, recipe_id, 'added', name, new_quantity=quantity, new_unit=unit)
    conn.commit()
    conn.close()
    return jsonify({'ingredient_id': ing_id, 'ingredient_name': name,
                    'quantity': quantity, 'unit': unit,
                    'ingredient_instruction': instruction}), 201


@app.route('/api/recipes/<int:recipe_id>/ingredients/<int:ingredient_id>', methods=['PATCH'])
def update_recipe_ingredient(recipe_id, ingredient_id):
    d = request.json or {}
    conn = get_db()
    _ensure_ingredient_editor_tables(conn)
    old = conn.execute(
        '''SELECT i.name_fi, ri.quantity, ri.unit FROM recipe_ingredients ri
           JOIN ingredients i ON i.id = ri.ingredient_id
           WHERE ri.recipe_id=? AND ri.ingredient_id=?''',
        (recipe_id, ingredient_id)).fetchone()
    if not old:
        conn.close()
        return jsonify({'error': 'Raaka-ainetta ei löydy tältä reseptiltä'}), 404

    new_quantity = old['quantity']
    if 'quantity' in d:
        try:
            new_quantity = round(float(d['quantity']), 3) if d['quantity'] not in (None, '') else None
        except (TypeError, ValueError):
            conn.close()
            return jsonify({'error': 'Virheellinen määrä'}), 400
    new_unit = (d.get('unit', old['unit']) or '').strip() or None
    new_instruction = None
    if 'ingredient_instruction' in d:
        new_instruction = (d.get('ingredient_instruction') or '').strip() or None

    fields, params = ['quantity = ?', 'unit = ?'], [new_quantity, new_unit]
    if 'ingredient_instruction' in d:
        fields.append('ingredient_instruction = ?')
        params.append(new_instruction)
    params += [recipe_id, ingredient_id]
    conn.execute(f'UPDATE recipe_ingredients SET {", ".join(fields)} '
                f'WHERE recipe_id=? AND ingredient_id=?', params)
    _log_ingredient_change(conn, recipe_id, 'updated', old['name_fi'],
                          old_quantity=old['quantity'], new_quantity=new_quantity,
                          old_unit=old['unit'], new_unit=new_unit)
    conn.commit()
    conn.close()
    return jsonify({'ingredient_id': ingredient_id, 'quantity': new_quantity, 'unit': new_unit})


@app.route('/api/recipes/<int:recipe_id>/ingredients/<int:ingredient_id>', methods=['DELETE'])
def delete_recipe_ingredient(recipe_id, ingredient_id):
    conn = get_db()
    _ensure_ingredient_editor_tables(conn)
    old = conn.execute(
        '''SELECT i.name_fi FROM recipe_ingredients ri JOIN ingredients i ON i.id = ri.ingredient_id
           WHERE ri.recipe_id=? AND ri.ingredient_id=?''',
        (recipe_id, ingredient_id)).fetchone()
    if not old:
        conn.close()
        return jsonify({'error': 'Raaka-ainetta ei löydy tältä reseptiltä'}), 404
    conn.execute('DELETE FROM recipe_ingredients WHERE recipe_id=? AND ingredient_id=?',
                (recipe_id, ingredient_id))
    _log_ingredient_change(conn, recipe_id, 'removed', old['name_fi'])
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/recipes/<int:recipe_id>/ingredients/<int:ingredient_id>/move', methods=['POST'])
def move_recipe_ingredient(recipe_id, ingredient_id):
    """Siirrä raaka-aine yhden askeleen ylös/alas nykyisessä
    näyttöjärjestyksessä (sort_order) — vaihtaa paikkaa vieressä olevan
    rivin kanssa. direction: 'up' tai 'down'."""
    d = request.json or {}
    direction = d.get('direction')
    if direction not in ('up', 'down'):
        return jsonify({'error': "direction pitää olla 'up' tai 'down'"}), 400

    conn = get_db()
    _ensure_ingredient_editor_tables(conn)
    rows = conn.execute(
        '''SELECT ri.ingredient_id, ri.sort_order
           FROM recipe_ingredients ri JOIN ingredients i ON i.id = ri.ingredient_id
           WHERE ri.recipe_id=? ORDER BY ri.sort_order, i.name_fi''', (recipe_id,)).fetchall()
    ids = [r['ingredient_id'] for r in rows]
    if ingredient_id not in ids:
        conn.close()
        return jsonify({'error': 'Raaka-ainetta ei löydy tältä reseptiltä'}), 404

    idx = ids.index(ingredient_id)
    swap_idx = idx - 1 if direction == 'up' else idx + 1
    if swap_idx < 0 or swap_idx >= len(ids):
        conn.close()
        return jsonify({'ok': True})  # jo listan reunassa — ei mitään tehtävää

    # Normalisoi järjestysnumerot 1..N ennen vaihtoa, jotta tasapelit
    # (esim. kaikki 0 ennen ensimmäistä käyttökertaa) eivät estä vaihtoa.
    for i, iid in enumerate(ids, start=1):
        conn.execute(
            'UPDATE recipe_ingredients SET sort_order=? WHERE recipe_id=? AND ingredient_id=?',
            (i, recipe_id, iid))
    ids[idx], ids[swap_idx] = ids[swap_idx], ids[idx]
    for i, iid in enumerate(ids, start=1):
        conn.execute(
            'UPDATE recipe_ingredients SET sort_order=? WHERE recipe_id=? AND ingredient_id=?',
            (i, recipe_id, iid))
    conn.commit()
    conn.close()
    return jsonify({'ok': True})


@app.route('/api/ingredients/list')
def list_ingredients():
    """Kaikki raaka-aineet käyttömäärineen — pohja selattavalle/haettavalle
    nimenkorjauslistalle (PoweResta-tuonneissa on paljon kirjoitusvirheitä ja
    -tuotenimen/työohjeen sekoittumia, joita on käytännössä mahdoton löytää
    ilman että näkee koko listan ja käyttömäärät kerralla)."""
    conn = get_db()
    rows = conn.execute(
        '''SELECT i.id, i.name_fi,
                  COUNT(DISTINCT ri.recipe_id) AS usage_count
           FROM ingredients i
           LEFT JOIN recipe_ingredients ri ON ri.ingredient_id = i.id
           GROUP BY i.id
           ORDER BY usage_count ASC, i.name_fi COLLATE NOCASE''').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@app.route('/api/ingredients/rename', methods=['POST'])
def rename_ingredient():
    """Nimeä raaka-aine uudelleen kaikkialla kertaheitolla. Reseptit eivät
    tallenna raaka-aineiden nimiä omiin JSON-blobeihinsa (se recipes.ingredients
    -sarake on käyttämätön kaikilla oikeilla tallennuspoluilla) — ne viittaavat
    yhteiseen ingredients-tauluun recipe_ingredients-liitostaulun kautta, joten
    yhden rivin nimen vaihto tässä taulussa riittää kaikille reseptejä
    käyttäville paikoille."""
    data = request.json or {}
    old_name = (data.get('old_name') or '').strip()
    new_name = (data.get('new_name') or '').strip()
    if not old_name or not new_name:
        return jsonify({'error': 'Anna sekä vanha että uusi nimi'}), 400
    if old_name == new_name:
        return jsonify({'error': 'Nimet ovat samat'}), 400

    conn = get_db()
    _ensure_ingredient_editor_tables(conn)
    old_row = conn.execute('SELECT id FROM ingredients WHERE name_fi = ?', (old_name,)).fetchone()
    if not old_row:
        conn.close()
        return jsonify({'error': f"Raaka-ainetta '{old_name}' ei löydy"}), 404
    old_id = old_row['id']

    affected = conn.execute(
        'SELECT COUNT(DISTINCT recipe_id) FROM recipe_ingredients WHERE ingredient_id = ?',
        (old_id,)).fetchone()[0]

    existing_new = conn.execute('SELECT id FROM ingredients WHERE name_fi = ?', (new_name,)).fetchone()
    if existing_new:
        # Kohdenimi on jo olemassa toisena raaka-aineena — yhdistä siihen sen
        # sijaan että rikottaisiin ingredients.name_fi:n UNIQUE-rajoitus.
        new_id = existing_new['id']
        # Jos jokin resepti käyttää jo MOLEMPIA (harvinainen, esim. vahingossa
        # kahteen kertaan lisätty), poista vanha rivi ettei recipe_ingredients-
        # taulun (recipe_id, ingredient_id) PRIMARY KEY rikkoudu.
        conflicting = conn.execute(
            '''SELECT recipe_id FROM recipe_ingredients
               WHERE ingredient_id = ? AND recipe_id IN
                     (SELECT recipe_id FROM recipe_ingredients WHERE ingredient_id = ?)''',
            (old_id, new_id)).fetchall()
        for row in conflicting:
            conn.execute('DELETE FROM recipe_ingredients WHERE recipe_id=? AND ingredient_id=?',
                        (row['recipe_id'], old_id))
        conn.execute('UPDATE recipe_ingredients SET ingredient_id = ? WHERE ingredient_id = ?',
                    (new_id, old_id))
        conn.execute('DELETE FROM ingredients WHERE id = ?', (old_id,))
    else:
        conn.execute('UPDATE ingredients SET name_fi = ? WHERE id = ?', (new_name, old_id))

    conn.commit()
    conn.close()
    return jsonify({'message': f"'{old_name}' → '{new_name}' päivitetty {affected} reseptiin",
                    'updated_recipes': affected})


# ============================================================
# INGREDIENT AUDIT: raportti resepteistä ilman raaka-aineita
# ============================================================
#
# HUOM lähdejaottelusta: reseptin alkuperäistä tuontikanavaa (PoweResta /
# OCR / verkkokaappaus) ei tallennettu ennen tätä ominaisuutta — se tieto
# hylättiin review_approve()-funktiossa. Jo olemassa olevia reseptejä ei
# voida siis jälkikäteen luotettavasti jakaa "poweresta" vs. "ocr" -kanaviin;
# ne näkyvät raportissa yhtenä "tuntematon (ennen lähdeseurantaa)" -ryhmänä.
# Uudet tuonnit tästä eteenpäin (PoweResta/OCR/verkkokaappaus/käsin liitetty/
# käyttäjän lisäämä/manuaalisesti kirjoitettu) merkitään oikein.

SOURCE_LABELS = {
    'poweresta': 'PoweResta-tuonti',
    'ocr': 'OCR-tuonti',
    'scraped': 'Verkkokaappaus',
    'manual_paste': 'Käsin liitetty',
    'manual': 'Kirjoitettu käsin',
    'user_contributed': 'Käyttäjän lisäämä',
    'legacy_tuntematon': 'Tuntematon (ennen lähdeseurantaa)',
    'tuntematon': 'Tuntematon',
}


def _audit_missing_query(source, category, season):
    _q = '''SELECT r.id, r.name_fi, r.source, r.dish_category, r.season, r.created_at
            FROM recipes r
            WHERE NOT EXISTS (SELECT 1 FROM recipe_ingredients ri WHERE ri.recipe_id = r.id)
              AND COALESCE(r.ingredients_audit_ok, 0) = 0'''
    params = []
    if source:
        _q += ' AND r.source = ?'
        params.append(source)
    if category:
        _q += ' AND r.dish_category = ?'
        params.append(category)
    if season:
        _q += ' AND r.season = ?'
        params.append(season)
    _q += ' ORDER BY r.source, r.name_fi'
    return _q, params


@app.route('/api/audit/ingredients/summary')
def audit_ingredients_summary():
    conn = get_db()
    _ensure_recipe_source_column(conn)
    _ensure_ingredient_editor_tables(conn)

    total = conn.execute('SELECT COUNT(*) FROM recipes').fetchone()[0]
    without_ing = conn.execute(
        '''SELECT COUNT(*) FROM recipes r
           WHERE NOT EXISTS (SELECT 1 FROM recipe_ingredients ri WHERE ri.recipe_id = r.id)
             AND COALESCE(r.ingredients_audit_ok, 0) = 0''').fetchone()[0]
    with_ing = total - without_ing

    breakdown = {}
    rows = conn.execute(
        '''SELECT source,
                  COUNT(*) AS total,
                  SUM(CASE WHEN EXISTS (SELECT 1 FROM recipe_ingredients ri
                                        WHERE ri.recipe_id = recipes.id)
                           OR COALESCE(ingredients_audit_ok, 0) = 1
                      THEN 1 ELSE 0 END) AS done
           FROM recipes GROUP BY source''').fetchall()
    for r in rows:
        src = r['source'] or 'tuntematon'
        done = r['done'] or 0
        tot = r['total']
        breakdown[src] = {
            'label': SOURCE_LABELS.get(src, src),
            'total': tot,
            'with_ingredients': done,
            'without_ingredients': tot - done,
            'percentage_complete': round(done / tot * 100) if tot else 0,
        }
    conn.close()

    return jsonify({
        'total_recipes': total,
        'recipes_with_ingredients': with_ing,
        'recipes_without_ingredients': without_ing,
        'breakdown_by_source': breakdown,
    })


@app.route('/api/audit/ingredients/missing')
def audit_missing_recipes():
    source = request.args.get('source', '')
    category = request.args.get('category', '')
    season = request.args.get('season', '')

    conn = get_db()
    _ensure_recipe_source_column(conn)
    _ensure_ingredient_editor_tables(conn)
    q, params = _audit_missing_query(source, category, season)
    rows = conn.execute(q, params).fetchall()
    conn.close()

    recipes = []
    for r in rows:
        d = dict(r)
        d['source_label'] = SOURCE_LABELS.get(d['source'] or 'tuntematon', d['source'])
        recipes.append(d)

    return jsonify({
        'filters_applied': {'source': source, 'category': category, 'season': season},
        'count': len(recipes),
        'recipes': recipes,
    })


@app.route('/api/audit/ingredients/mark-complete/<int:recipe_id>', methods=['POST'])
def audit_mark_complete(recipe_id):
    d = request.json or {}
    notes = (d.get('notes') or '').strip() or 'Tarkistettu — ei tarvitse raaka-ainelistaa'
    conn = get_db()
    _ensure_recipe_source_column(conn)
    _ensure_ingredient_editor_tables(conn)
    recipe = conn.execute('SELECT name_fi FROM recipes WHERE id=?', (recipe_id,)).fetchone()
    if not recipe:
        conn.close()
        return jsonify({'error': 'Reseptiä ei löydy'}), 404
    conn.execute('UPDATE recipes SET ingredients_audit_ok=1 WHERE id=?', (recipe_id,))
    conn.execute(
        '''INSERT INTO ingredient_changes
           (recipe_id, changed_by_user_id, changed_by_username, change_type,
            ingredient_name, notes)
           VALUES (?,?,?,'audit_checked',?,?)''',
        (recipe_id, None, 'admin',
         '(koko reseptin tarkistus)', notes))
    conn.commit()
    conn.close()
    return jsonify({'ok': True, 'message': f"'{recipe['name_fi']}' merkitty tarkistetuksi"})


@app.route('/api/audit/ingredients/progress')
def audit_progress():
    conn = get_db()
    _ensure_recipe_source_column(conn)
    _ensure_ingredient_editor_tables(conn)

    to_audit = conn.execute(
        '''SELECT COUNT(*) FROM recipes r
           WHERE NOT EXISTS (SELECT 1 FROM recipe_ingredients ri WHERE ri.recipe_id = r.id)
             AND COALESCE(r.ingredients_audit_ok, 0) = 0''').fetchone()[0]
    today = conn.execute(
        '''SELECT COUNT(*) FROM ingredient_changes
           WHERE change_type='audit_checked' AND date(timestamp) = date('now')''').fetchone()[0]
    this_week = conn.execute(
        '''SELECT COUNT(*) FROM ingredient_changes
           WHERE change_type='audit_checked' AND date(timestamp) >= date('now', '-6 days')'''
    ).fetchone()[0]
    last = conn.execute(
        '''SELECT changed_by_username, timestamp FROM ingredient_changes
           WHERE change_type='audit_checked' ORDER BY id DESC LIMIT 1''').fetchone()
    conn.close()

    return jsonify({
        'recipes_to_audit': to_audit,
        'recipes_audited_today': today,
        'recipes_audited_this_week': this_week,
        'last_audit_by': last['changed_by_username'] if last else None,
        'last_audit_time': last['timestamp'] if last else None,
    })


@app.route('/api/audit/ingredients/export')
def audit_export():
    import csv
    import io

    source = request.args.get('source', '')
    category = request.args.get('category', '')
    season = request.args.get('season', '')

    conn = get_db()
    _ensure_recipe_source_column(conn)
    _ensure_ingredient_editor_tables(conn)
    q, params = _audit_missing_query(source, category, season)
    rows = conn.execute(q, params).fetchall()
    conn.close()

    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(['ID', 'Reseptin nimi', 'Lähde', 'Kategoria', 'Sesonki', 'Luotu'])
    for r in rows:
        writer.writerow([r['id'], r['name_fi'], SOURCE_LABELS.get(r['source'] or 'tuntematon', r['source']),
                         r['dish_category'], r['season'], r['created_at']])

    data = buf.getvalue().encode('utf-8-sig')  # BOM: Excel avaa ä/ö oikein
    stamp = datetime.now().strftime('%Y%m%d')
    return send_file(io.BytesIO(data), mimetype='text/csv', as_attachment=True,
                     download_name=f'puuttuvat_raaka_aineet_{stamp}.csv')


# ============================================================
# REVIEW QUEUE: scraped recipes awaiting approval
# ============================================================

REVIEW_FILE = os.path.join(_DATA_DIR, 'recipes_for_review.json')
if not os.path.exists(REVIEW_FILE):
    _legacy_review = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'recipes_for_review.json')
    if os.path.exists(_legacy_review) and os.path.abspath(_legacy_review) != os.path.abspath(REVIEW_FILE):
        import shutil as _shutil
        _shutil.copy2(_legacy_review, REVIEW_FILE)


@app.route('/api/translate', methods=['POST'])
def translate_text():
    """Kääntää lyhyen englanninkielisen tekstin suomeksi paikallisesti
    (ei verkkoyhteyttä, ei ulkoista API:a) — tarkoitettu käsin lisättyjen
    tai kerättyjen englanninkielisten reseptien nimien/raaka-aineiden/
    ohjeiden esikäännökseksi. Käyttäjän pitää aina tarkistaa tulos ennen
    hyväksymistä, konekäännös ei ole täydellinen etenkään yksiköiden osalta."""
    data = request.get_json(force=True) or {}
    text = data.get('text', '')
    translated, error = translate_en_to_fi(text)
    if error:
        return jsonify({'error': error}), 500
    return jsonify({'translated': translated})


@app.route('/api/review')
def review_list():
    if not os.path.exists(REVIEW_FILE):
        return jsonify([])
    with open(REVIEW_FILE, encoding='utf-8') as f:
        return jsonify(json.load(f))


@app.route('/api/review', methods=['POST'])
def review_save():
    """Save the (modified) review list back to disk."""
    with open(REVIEW_FILE, 'w', encoding='utf-8') as f:
        json.dump(request.json, f, ensure_ascii=False, indent=2)
    return jsonify({'message': 'Tallennettu'})


approve_state = {
    'running': False, 'total': 0, 'done': 0,
    'result': None, 'error': None
}
approve_lock = threading.Lock()


def _approve_worker(recipes):
    """Background worker: upload approved review-queue recipes into the database."""
    try:
        db = MealPlanDB(DB_PATH)
        ok, updated, skipped, merged = 0, 0, [], []
        for r in recipes:
            if r.get('deleted'):
                with approve_lock:
                    approve_state['done'] += 1
                continue
            if not r.get('name_fi'):
                skipped.append(r.get('name_fi') or '(nimetön)')
                if r.get('name_fi'):
                    # resepti saattaa olla jo tietokannassa (esim. kausi/kategoria
                    # puuttuu tuonnista) — liitä raaka-aineet siihen silti
                    conn = get_db()
                    row = conn.execute('SELECT id FROM recipes WHERE name_fi=?',
                                       (r['name_fi'],)).fetchone()
                    conn.close()
                    _save_recipe_ingredients(row['id'] if row else None,
                                             r.get('ingredients_struct') or [])
                with approve_lock:
                    approve_state['done'] += 1
                continue
            # Keep otherwise valid recipes from being stuck in the review queue
            # merely because automatic classification was inconclusive.
            if r.get('season') not in SEASONS:
                r['season'] = 'kaikki'
            if r.get('dish_category') not in CATEGORIES:
                r['dish_category'] = 'kasvis'
            if r.get('recipe_type') not in RECIPE_TYPES:
                # Ei eksplisiittistä valintaa Tarkistusjonossa — arvaa nimestä
                # samalla logiikalla jota generaattori käytti ennen recipe_type-
                # saraketta, jottei tuonti jää käsittelemättömäksi tämän takia.
                name_lower = r['name_fi'].lower()
                if 'salaa' in name_lower:
                    r['recipe_type'] = 'salaatti'
                elif 'keitto' in name_lower:
                    r['recipe_type'] = 'keitto'
                else:
                    r['recipe_type'] = 'pääruoka'
            rid = db.add_recipe(
                name_fi=r['name_fi'], season=r['season'],
                meal_type=r.get('meal_type', 'lounas'),
                dish_category=r['dish_category'],
                recipe_type=r['recipe_type'],
                prep_time_min=r.get('prep_time_min'),
                source_url=r.get('source_url'),
                notes=r.get('notes', ''),
                servings=r.get('servings'),
                instructions=r.get('instructions_raw', '')
            )
            if rid:
                ok += 1
                conn = get_db()
                _ensure_recipe_source_column(conn)
                _set_recipe_source(conn, rid, _classify_import_source(
                    r.get('source_site', ''), r.get('source_url', '')))
                conn.commit()
                conn.close()
            else:
                # hae olemassa oleva resepti, jotta raaka-aineet voidaan silti liittää
                conn = get_db()
                row = conn.execute('SELECT id FROM recipes WHERE name_fi=?',
                                   (r['name_fi'],)).fetchone()
                rid = row['id'] if row else None
                existing_has_ingredients = bool(rid) and conn.execute(
                    'SELECT COUNT(*) FROM recipe_ingredients WHERE recipe_id=?',
                    (rid,)).fetchone()[0] > 0
                conn.close()
                ingredients_present = bool(r.get('ingredients_struct'))
                if r.get('review_action') == 'updated':
                    updated += 1
                elif rid and not existing_has_ingredients and ingredients_present:
                    # Olemassa oleva resepti ei sisällä vielä raaka-aineita — tuonnin
                    # raaka-aineet täydentävät sen automaattisesti alla olevassa
                    # _save_recipe_ingredients-kutsussa sen sijaan, että jäisivät ohi.
                    merged.append(r['name_fi'])
                else:
                    skipped.append(r['name_fi'] + ' (jo olemassa)')
            ingredients = r.get('ingredients_struct') or []
            if r.get('review_action') == 'updated' and rid:
                _replace_recipe_ingredients(rid, ingredients)
            else:
                _save_recipe_ingredients(rid, ingredients)
            with approve_lock:
                approve_state['done'] += 1
        # clear the queue of uploaded items
        remaining = [r for r in recipes if r.get('deleted') or not r.get('name_fi')]
        remaining = [r for r in remaining if not r.get('deleted')]
        with open(REVIEW_FILE, 'w', encoding='utf-8') as f:
            json.dump(remaining, f, ensure_ascii=False, indent=2)
        with approve_lock:
            approve_state['result'] = {'uploaded': ok, 'updated': updated, 'skipped': skipped, 'merged': merged}
    except Exception as e:
        with approve_lock:
            approve_state['error'] = str(e)
    finally:
        with approve_lock:
            approve_state['running'] = False


@app.route('/api/review/approve', methods=['POST'])
def review_approve():
    """Start uploading approved review-queue recipes into the database in the background."""
    recipes = request.json or []
    with approve_lock:
        if approve_state['running']:
            return jsonify({'error': 'Vienti on jo käynnissä — odota sen valmistumista'}), 409
        approve_state.update({'running': True, 'total': len(recipes), 'done': 0,
                              'result': None, 'error': None})
    threading.Thread(target=_approve_worker, args=(recipes,), daemon=True).start()
    return jsonify({'message': f'Vienti aloitettu: {len(recipes)} reseptiä', 'total': len(recipes)})


@app.route('/api/review/approve/status')
def review_approve_status():
    with approve_lock:
        return jsonify(dict(approve_state))


# ============================================================
# MEAL PLANS: generate / view / modify / export
# ============================================================

FACILITIES = {'kesti': 'Kesti', 'hoiva': 'Hoiva', 'kymenkartano': 'Kymenkartano'}
# Kesti: ma-pe (5). Hoiva ja Kymenkartano: ma-su (7) — molemmat samasta
# jaetusta reseptipoolista, mutta kumpikin oma erikseen generoitu suunnitelmansa.
FACILITY_DAYS_PER_WEEK = {'kesti': 5, 'hoiva': 7, 'kymenkartano': 7}


@app.route('/api/plans')
def list_plans():
    conn = get_db()
    rows = conn.execute('SELECT id, name, season, num_weeks, facility, created_at FROM meal_plans '
                        'ORDER BY created_at DESC').fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


def _plan_season_distribution(plan_id):
    """Kuinka moni suunnitelman viikoista osuu kullekin kaudelle (kalenteripohjaisesti)."""
    counts = {'talvi': 0, 'kevät': 0, 'kesä': 0, 'syksy': 0}
    for season, start_week, end_week in MealPlanGenerator.YEAR_SEGMENTS:
        counts[season] += end_week - start_week + 1
    return counts


@app.route('/api/plans', methods=['POST'])
def create_plan():
    data = request.json or {}
    season = data.get('season')
    num_weeks = int(data.get('num_weeks', 52))
    only_new = bool(data.get('only_new_recipes'))
    facility = data.get('facility') if data.get('facility') in FACILITIES else 'kesti'

    if facility == 'hoiva':
        # Hoiva ei ole enää itsenäisesti generoitu toimipiste: sen ma-pe on
        # AINA tasan sama kuin nykyisen Kesti-vuosisuunnitelman (vahvistettu
        # käyttäjältä) — vain la-su generoidaan ja tallennetaan Hoivalle
        # itselleen. season/num_weeks-parametrit eivät koske Hoivaa lainkaan.
        conn = get_db()
        kesti_plan = conn.execute(
            "SELECT id, name FROM meal_plans WHERE facility = 'kesti' AND season = 'vuosi' "
            "AND name NOT LIKE '%(arkistoitu %' ORDER BY created_at DESC LIMIT 1").fetchone()
        if not kesti_plan:
            conn.close()
            return jsonify({'error': 'Kestin vuosisuunnitelmaa ei löydy — luo se ensin, '
                                     'Hoiva peilaa sitä.'}), 400
        existing_hoiva = conn.execute(
            "SELECT id, name FROM meal_plans WHERE facility = 'hoiva' "
            "AND name NOT LIKE '%(arkistoitu %' ORDER BY created_at DESC LIMIT 1").fetchone()
        if existing_hoiva and not data.get('replace'):
            conn.close()
            return jsonify({
                'error': 'Hoiva: la-su-suunnitelma on jo olemassa. Lähetä replace: true korvataksesi sen.',
                'existing_plan_id': existing_hoiva['id'],
                'existing_plan_name': existing_hoiva['name'],
            }), 409
        if existing_hoiva:
            archived_at = datetime.now().strftime('%Y-%m-%d %H:%M')
            conn.execute('UPDATE meal_plans SET name = ? WHERE id = ?',
                        (f"{existing_hoiva['name']} (arkistoitu {archived_at})", existing_hoiva['id']))
            conn.commit()
        conn.close()

        hoiva_generator = MealPlanGenerator(DB_PATH, facility='hoiva')
        plan_id = hoiva_generator.generate_weekend_extension(kesti_plan['id'], only_new=only_new)
        if not plan_id:
            return jsonify({'error': 'Hoivan la-su-suunnitelman luonti epäonnistui'}), 500
        return jsonify({
            'success': True,
            'meal_plan_id': plan_id,
            'mirrors_plan_id': kesti_plan['id'],
            'message': f"Hoiva luotu — ma-pe on aina sama kuin \"{kesti_plan['name']}\", "
                       'la-su generoitu erikseen. Mene Selaa & Muokkaa -välilehteen.',
        })

    generator = MealPlanGenerator(DB_PATH, days_per_week=FACILITY_DAYS_PER_WEEK[facility], facility=facility)

    if season == 'vuosi':
        # Älä hiljaa luo toista vuosisuunnitelmaa olemassa olevan rinnalle —
        # esihenkilön pitää joko korvata vanha (replace=true) tai poistaa se itse ensin.
        # Arkistoidut (korvatut) vuosisuunnitelmat eivät lasketa "olemassa olevaksi" tässä.
        # Jokaisella toimipisteellä (kesti/hoiva/kymenkartano) on oma erillinen
        # vuosisuunnitelmansa, joten tarkistus on rajattu samaan toimipisteeseen.
        if not data.get('replace'):
            conn = get_db()
            existing = conn.execute(
                "SELECT id, name, created_at FROM meal_plans WHERE season = 'vuosi' "
                "AND facility = ? AND name NOT LIKE '%(arkistoitu %' "
                "ORDER BY created_at DESC LIMIT 1", (facility,)).fetchone()
            conn.close()
            if existing:
                return jsonify({
                    'error': f'{FACILITIES[facility]}: vuosisuunnitelma on jo olemassa. Lähetä '
                             'replace: true korvataksesi sen, tai poista se ensin.',
                    'existing_plan_id': existing['id'],
                    'existing_plan_name': existing['name'],
                }), 409

        # 52 weeks split into 4 seasonal themes by calendar
        shortages = []
        for s in ('talvi', 'kevät', 'kesä', 'syksy'):
            n = len(generator._get_season_pool(s, only_new=only_new))
            if n < 25:
                shortages.append(f'{SEASONS[s]}: {n}')
        if shortages:
            scope = 'uusista (käsin lisätyistä/muokatuista) resepteistä' if only_new else 'reseptejä'
            return jsonify({'error': f'Liian vähän {scope} kausille — ' + ', '.join(shortages) +
                                     '. Jokaiselle kaudelle tarvitaan vähintään 25 reseptiä '
                                     '(ympärivuotiset lasketaan mukaan joka kauteen).'}), 400

        if data.get('replace'):
            # Arkistoi vanhat vuosisuunnitelmat sen sijaan että poistettaisiin ne —
            # ne pysyvät Ruokalistat-listalla katselua/vertailua/myöhempää poistoa varten.
            conn = get_db()
            old_rows = conn.execute(
                "SELECT id, name FROM meal_plans WHERE season = 'vuosi' AND facility = ? "
                "AND name NOT LIKE '%(arkistoitu %'", (facility,)).fetchall()
            archived_at = datetime.now().strftime('%Y-%m-%d %H:%M')
            for row in old_rows:
                conn.execute('UPDATE meal_plans SET name = ? WHERE id = ?',
                             (f"{row['name']} (arkistoitu {archived_at})", row['id']))
            conn.commit()
            conn.close()

        plan_id = generator.generate_year_plan(only_new=only_new, facility=facility)
        if not plan_id:
            return jsonify({'error': 'Ruokalistan luonti epäonnistui'}), 500

        if facility == 'kesti' and data.get('replace') and old_rows:
            # Hoiva peilaa Kestin vuosisuunnitelmaa — kun se korvataan, minkä
            # tahansa elävän Hoiva-suunnitelman peili pitää osoittaa uuteen
            # (muuten se jäisi osoittamaan juuri arkistoituun vanhaan).
            conn = get_db()
            old_ids = [r['id'] for r in old_rows]
            placeholders = ','.join('?' * len(old_ids))
            conn.execute(
                f"UPDATE meal_plans SET mirrors_plan_id = ? "
                f"WHERE facility = 'hoiva' AND mirrors_plan_id IN ({placeholders}) "
                f"AND name NOT LIKE '%(arkistoitu %'",
                [plan_id] + old_ids)
            conn.commit()
            conn.close()

        stats = generator.get_meal_plan_stats(plan_id)
        protein_total = sum(stats['category_distribution'].values()) or 1

        conn = get_db()
        salad_rows = conn.execute(
            '''SELECT r.name_fi, COUNT(*) as n FROM meal_plan_days d
               JOIN recipes r ON d.recipe_id = r.id
               WHERE d.meal_plan_id = ? AND d.meal_type = 'salaatti'
               GROUP BY r.name_fi ORDER BY n DESC, r.name_fi''',
            (plan_id,)).fetchall()
        soup_rows = conn.execute(
            '''SELECT r.name_fi, COUNT(*) as n FROM meal_plan_days d
               JOIN recipes r ON d.recipe_id = r.id
               WHERE d.meal_plan_id = ? AND d.meal_type = 'keitto'
               GROUP BY r.name_fi ORDER BY n DESC, r.name_fi''',
            (plan_id,)).fetchall()
        conn.close()

        return jsonify({
            'success': True,
            'meal_plan_id': plan_id,
            'weeks_generated': 52,
            'total_meals': stats['total_meals'],
            'recipes_used': stats['unique_recipes'],
            'season_distribution': _plan_season_distribution(plan_id),
            'protein_distribution': {
                cat: f'~{round(100 * n / protein_total)}%'
                for cat, n in stats['category_distribution'].items()
            },
            'salads_in_rotation': len(salad_rows),
            'salad_usage': {row['name_fi']: row['n'] for row in salad_rows},
            'soups_in_rotation': len(soup_rows),
            'soup_usage': {row['name_fi']: row['n'] for row in soup_rows},
            'message': '52 viikon ruokalista luotu — jokaisella arkipäivällä oma keitto, '
                       'pääruoka ja salaatti. Mene Selaa & Muokkaa -välilehteen muokkaaksesi.',
        })

    if season not in SEASONS or season == 'kaikki':
        return jsonify({'error': 'Valitse kausi'}), 400
    count = len(generator._get_season_pool(season, only_new=only_new))
    if count < 25:
        scope = 'uusia (käsin lisättyjä/muokattuja) reseptejä' if only_new else 'reseptiä'
        return jsonify({'error': f'Kaudella "{SEASONS[season]}" on vain {count} {scope} '
                                 f'(ml. ympärivuotiset). Tarvitaan vähintään 25.'}), 400
    plan_id = generator.generate_meal_plan(season, num_weeks=num_weeks, only_new=only_new, facility=facility)
    if not plan_id:
        return jsonify({'error': 'Ruokalistan luonti epäonnistui'}), 500
    stats = generator.get_meal_plan_stats(plan_id)
    return jsonify({'id': plan_id, 'stats': {
        'total_meals': stats['total_meals'],
        'unique_recipes': stats['unique_recipes'],
        'category_distribution': stats['category_distribution'],
    }})


MEAL_ROLE_BY_TYPE = {'keitto': 'soup', 'salaatti': 'salad', 'lounas2': 'main2', 'lounas3': 'main3',
                      'kastike': 'sauce', 'kasvislisäke': 'veg_side', 'energialisäke': 'energy_side'}  # anything else ('lounas') -> main
ACCOMPANIMENT_ROLES = ('sauce', 'veg_side', 'energy_side')
ACCOMPANIMENT_TYPE_BY_ROLE = {'sauce': 'kastike', 'veg_side': 'kasvislisäke', 'energy_side': 'energialisäke'}
WEEKDAY_LABELS = ['Ma', 'Ti', 'Ke', 'To', 'Pe', 'La', 'Su']  # kesti käyttää vain 0-4, hoiva/kymenkartano 0-6
# Mikä viikonpäivä (0=ma..6=su) saa jälkiruokarivin, toimipisteittäin.
# Kesti: perjantai (oma pohja ei näytä sitä, mutta rivi tallennetaan silti).
# Hoiva: sunnuntai — vahvistettu asukas-pohjan JÄLKIRUOKA-rivistä.
DESSERT_DAY_BY_FACILITY = {'kesti': 4, 'hoiva': 6, 'kymenkartano': 6}


def _mirrored_day_error(conn, plan_id, day_of_week):
    """Hoiva's ma-pe (day_of_week 0-4) is a live mirror of its linked Kesti
    plan — those rows don't exist on the Hoiva plan itself, so editing them
    there would either fail silently or (worse) create a rogue row that
    breaks the mirror. Returns a jsonify()able error dict if this edit
    should be blocked, else None."""
    if day_of_week >= 5:
        return None
    plan = conn.execute('SELECT mirrors_plan_id FROM meal_plans WHERE id = ?', (plan_id,)).fetchone()
    if plan and plan['mirrors_plan_id']:
        return {'error': 'Tämä päivä on aina sama kuin Kestin ruokalistalla — muokkaa sitä '
                         'Kestin suunnitelman kautta.', 'mirrors_plan_id': plan['mirrors_plan_id']}
    return None


def _effective_meal_plan_days(conn, plan_id, week=None):
    """Return this plan's day-by-day meal_plan_days rows (week_number,
    day_of_week, recipe_id, meal_type, dessert_text), transparently
    resolving a mirrored facility (Hoiva, confirmed to always equal Kesti's
    plan Mon-Fri) so callers never need to know about the mirror: ma-pe
    come from the linked Kesti plan, la-su from the plan's own rows.
    Non-mirrored plans (kesti, kymenkartano) just return their own rows."""
    plan = conn.execute('SELECT mirrors_plan_id FROM meal_plans WHERE id = ?', (plan_id,)).fetchone()
    mirrors = plan['mirrors_plan_id'] if plan else None
    week_clause = 'AND week_number = ?' if week is not None else ''
    week_params = (week,) if week is not None else ()

    own_rows = conn.execute(
        f'''SELECT week_number, day_of_week, recipe_id, meal_type, dessert_text
            FROM meal_plan_days WHERE meal_plan_id = ? {week_clause}''',
        (plan_id,) + week_params).fetchall()
    if not mirrors:
        return sorted(own_rows, key=lambda r: (r['week_number'], r['day_of_week']))

    accompaniment_types = ','.join('?' * len(ACCOMPANIMENT_TYPE_BY_ROLE))
    mirrored_rows = conn.execute(
        f'''SELECT week_number, day_of_week, recipe_id, meal_type, dessert_text
            FROM meal_plan_days WHERE meal_plan_id = ? AND day_of_week < 5
                AND meal_type NOT IN ({accompaniment_types}) {week_clause}''',
        (mirrors,) + tuple(ACCOMPANIMENT_TYPE_BY_ROLE.values()) + week_params).fetchall()
    own_weekend = [r for r in own_rows if r['day_of_week'] >= 5]
    return sorted(list(mirrored_rows) + own_weekend, key=lambda r: (r['week_number'], r['day_of_week']))


@app.route('/api/plans/<int:plan_id>')
def get_plan(plan_id):
    conn = get_db()
    plan = conn.execute('SELECT * FROM meal_plans WHERE id = ?', (plan_id,)).fetchone()
    if not plan:
        conn.close()
        return jsonify({'error': 'Ruokalistaa ei löydy'}), 404
    day_rows = _effective_meal_plan_days(conn, plan_id)
    recipe_ids = list({r['recipe_id'] for r in day_rows})
    recipes_by_id = {}
    if recipe_ids:
        placeholders = ','.join('?' * len(recipe_ids))
        recipes_by_id = {rr['id']: rr for rr in conn.execute(
            f'''SELECT id, name_fi, dish_category,
                       (SELECT COUNT(*) FROM recipe_ingredients ri WHERE ri.recipe_id = recipes.id) AS ingredient_count
                FROM recipes WHERE id IN ({placeholders})''', recipe_ids).fetchall()}
    conn.close()

    weeks_grouped = defaultdict(lambda: defaultdict(
        lambda: {'day': None, 'main': None, 'main2': None, 'main3': None, 'soup': None, 'salad': None,
                 'sauce': None, 'veg_side': None, 'energy_side': None, 'dessert_text': None}))
    for r in day_rows:
        day_obj = weeks_grouped[r['week_number']][r['day_of_week']]
        day_obj['day'] = r['day_of_week']
        role = MEAL_ROLE_BY_TYPE.get(r['meal_type'], 'main')
        if r['recipe_id'] is None:
            # Slot was cleared by hand (see /clear_meal) — keep the row so it
            # still occupies its place in the day and stays a drop target,
            # but don't show it as a broken reference like a dangling '?'.
            day_obj[role] = {'recipe_id': None, 'name': None, 'category': None,
                             'meal_type': r['meal_type'], 'has_ingredients': True}
        else:
            rr = recipes_by_id.get(r['recipe_id'])
            day_obj[role] = {'recipe_id': r['recipe_id'], 'name': rr['name_fi'] if rr else '?',
                             'category': rr['dish_category'] if rr else None, 'meal_type': r['meal_type'],
                             'has_ingredients': (rr['ingredient_count'] > 0) if rr else False}
        if role == 'main':
            day_obj['dessert_text'] = r['dessert_text']

    weeks = {week_num: sorted(days.values(), key=lambda d: d['day'])
             for week_num, days in weeks_grouped.items()}
    return jsonify({'id': plan['id'], 'name': plan['name'], 'season': plan['season'],
                    'num_weeks': plan['num_weeks'], 'facility': plan['facility'] or 'kesti',
                    'mirrors_plan_id': plan['mirrors_plan_id'],
                    'weeks': weeks})


@app.route('/api/plans/<int:plan_id1>/compare/<int:plan_id2>')
def compare_plans(plan_id1, plan_id2):
    """Vertaa kahden ruokalistan tiettyä viikkoa (oletus vko 1) päivä kerrallaan."""
    try:
        week = _int_arg(request.args, 'week', 1)
    except _BadIntArg as e:
        return jsonify({'error': e.message}), 400
    conn = get_db()
    plans = {}
    for pid in (plan_id1, plan_id2):
        plan = conn.execute('SELECT id, name FROM meal_plans WHERE id = ?', (pid,)).fetchone()
        if not plan:
            conn.close()
            return jsonify({'error': f'Ruokalistaa {pid} ei löydy'}), 404
        plans[pid] = plan

    rows = conn.execute(
        '''SELECT d.meal_plan_id, d.day_of_week, d.meal_type, r.name_fi
           FROM meal_plan_days d JOIN recipes r ON d.recipe_id = r.id
           WHERE d.meal_plan_id IN (?, ?) AND d.week_number = ?
           ORDER BY d.day_of_week''', (plan_id1, plan_id2, week)).fetchall()
    conn.close()

    by_day = defaultdict(lambda: {plan_id1: {}, plan_id2: {}})
    for r in rows:
        role = MEAL_ROLE_BY_TYPE.get(r['meal_type'], 'main')
        by_day[r['day_of_week']][r['meal_plan_id']][role] = r['name_fi']

    comparison = []
    changes = 0
    for day in sorted(by_day.keys()):
        a, b = by_day[day][plan_id1], by_day[day][plan_id2]
        same = a == b
        if not same:
            changes += 1
        comparison.append({
            'day_of_week': WEEKDAY_LABELS[day] if 0 <= day < len(WEEKDAY_LABELS) else str(day),
            'plan1': {'soup': a.get('soup'), 'main': a.get('main'), 'salad': a.get('salad')},
            'plan2': {'soup': b.get('soup'), 'main': b.get('main'), 'salad': b.get('salad')},
            'same': same,
        })

    return jsonify({
        'week': week,
        'plan1': {'id': plan_id1, 'name': plans[plan_id1]['name']},
        'plan2': {'id': plan_id2, 'name': plans[plan_id2]['name']},
        'comparison': comparison,
        'changes': changes,
    })


@app.route('/api/plans/<int:plan_id>', methods=['DELETE'])
def delete_plan(plan_id):
    conn = get_db()
    conn.execute('DELETE FROM meal_plan_days WHERE meal_plan_id = ?', (plan_id,))
    conn.execute('DELETE FROM meal_plans WHERE id = ?', (plan_id,))
    conn.commit()
    conn.close()
    return jsonify({'message': 'Ruokalista poistettu'})


@app.route('/api/plans/<int:plan_id>/change_meal', methods=['POST'])
def change_meal(plan_id):
    """Change one meal slot - the 'short notice' feature."""
    data = request.json or {}
    try:
        week_number = _int_arg(data, 'week')
        day_of_week = _int_arg(data, 'slot')
        new_recipe_id = _int_arg(data, 'new_recipe_id')
    except _BadIntArg as e:
        return jsonify({'error': e.message}), 400
    conn = get_db()
    mirror_err = _mirrored_day_error(conn, plan_id, day_of_week)
    conn.close()
    if mirror_err:
        return jsonify(mirror_err), 400
    modifier = MealModifier(DB_PATH)
    ok = modifier.change_meal(
        meal_plan_id=plan_id,
        week_number=week_number,
        day_of_week=day_of_week,
        meal_type=data.get('meal_type', 'lounas'),
        new_recipe_id=new_recipe_id,
        reason=data.get('reason', 'Manuaalinen vaihto'),
        modified_by=data.get('modified_by', 'käyttäjä')
    )
    if not ok:
        return jsonify({'error': 'Vaihto epäonnistui'}), 400
    return jsonify({'message': 'Ateria vaihdettu'})


@app.route('/api/plans/<int:plan_id>/clear_meal', methods=['POST'])
def clear_meal(plan_id):
    """Remove the recipe from a required slot (soup/main/salad), leaving it
    empty rather than deleting the day's row — the slot stays a valid
    drag/click target to fill in again later."""
    data = request.json or {}
    try:
        week_number = _int_arg(data, 'week')
        day_of_week = _int_arg(data, 'slot')
    except _BadIntArg as e:
        return jsonify({'error': e.message}), 400
    conn = get_db()
    mirror_err = _mirrored_day_error(conn, plan_id, day_of_week)
    conn.close()
    if mirror_err:
        return jsonify(mirror_err), 400
    modifier = MealModifier(DB_PATH)
    ok = modifier.clear_meal(
        meal_plan_id=plan_id,
        week_number=week_number,
        day_of_week=day_of_week,
        meal_type=data.get('meal_type', 'lounas'),
        reason=data.get('reason', 'Poistettu'),
        modified_by=data.get('modified_by', 'käyttäjä')
    )
    if not ok:
        return jsonify({'error': 'Poisto epäonnistui'}), 400
    return jsonify({'message': 'Ateria poistettu'})


@app.route('/api/plans/<int:plan_id>/dessert', methods=['POST'])
def update_dessert(plan_id):
    """Tallenna vapaamuotoinen jälkiruokateksti yhdelle päivälle (perjantaille,
    day_of_week=4, mutta toimii mille tahansa päivälle jos joskus tarvitaan
    muillekin). Tallennetaan päivän 'lounas'-riville, koska jokaisella
    päivällä on aina tasan yksi sellainen."""
    data = request.json or {}
    try:
        week = _int_arg(data, 'week')
        day_of_week = _int_arg(data, 'day_of_week')
    except _BadIntArg as e:
        return jsonify({'error': e.message}), 400
    dessert_text = (data.get('dessert_text') or '').strip()

    conn = get_db()
    mirror_err = _mirrored_day_error(conn, plan_id, day_of_week)
    if mirror_err:
        conn.close()
        return jsonify(mirror_err), 400
    cur = conn.execute(
        '''UPDATE meal_plan_days SET dessert_text = ?
           WHERE meal_plan_id = ? AND week_number = ? AND day_of_week = ? AND meal_type = 'lounas' ''',
        (dessert_text, plan_id, week, day_of_week))
    conn.commit()
    updated = cur.rowcount
    conn.close()
    if not updated:
        return jsonify({'error': 'Päivää ei löytynyt tältä ruokalistalta'}), 404
    return jsonify({'message': 'Jälkiruoka tallennettu'})


def _update_extra_main(plan_id, meal_type, label, require_kesti=False):
    """Shared logic for an optional extra slot on one day (meal_type
    'lounas2'/'lounas3', or one of the Kesti-only accompaniment types
    'kastike'/'kasvislisäke'/'energialisäke') — not something the generator
    ever assigns automatically, only added/removed by hand. Stored as its
    own row alongside the always-present 'lounas' row. recipe_id=None (or
    missing) removes it. require_kesti=True rejects the call outright if
    the plan isn't Kesti's (these accompaniments must never exist on
    Hoiva's/Kymenkartano's own rows — Hoiva only ever sees them via the
    mirror of Kesti's data)."""
    data = request.json or {}
    try:
        week = _int_arg(data, 'week')
        day_of_week = _int_arg(data, 'day_of_week')
    except _BadIntArg as e:
        return jsonify({'error': e.message}), 400
    recipe_id = data.get('recipe_id')

    conn = get_db()
    if require_kesti:
        plan = conn.execute('SELECT facility FROM meal_plans WHERE id = ?', (plan_id,)).fetchone()
        if not plan or (plan['facility'] or 'kesti') != 'kesti':
            conn.close()
            return jsonify({'error': f'{label} koskee vain Kestiä'}), 400
    mirror_err = _mirrored_day_error(conn, plan_id, day_of_week)
    if mirror_err:
        conn.close()
        return jsonify(mirror_err), 400
    if recipe_id in (None, ''):
        cur = conn.execute(
            f'''DELETE FROM meal_plan_days
                WHERE meal_plan_id = ? AND week_number = ? AND day_of_week = ? AND meal_type = '{meal_type}' ''',
            (plan_id, week, day_of_week))
        conn.commit()
        conn.close()
        return jsonify({'message': f'{label} poistettu', 'removed': cur.rowcount > 0})

    try:
        recipe_id = int(recipe_id)
    except (TypeError, ValueError):
        conn.close()
        return jsonify({'error': 'Virheellinen resepti'}), 400
    if not conn.execute('SELECT id FROM recipes WHERE id = ?', (recipe_id,)).fetchone():
        conn.close()
        return jsonify({'error': 'Reseptiä ei löydy'}), 404

    existing = conn.execute(
        f'''SELECT id FROM meal_plan_days
            WHERE meal_plan_id = ? AND week_number = ? AND day_of_week = ? AND meal_type = '{meal_type}' ''',
        (plan_id, week, day_of_week)).fetchone()
    if existing:
        conn.execute('UPDATE meal_plan_days SET recipe_id = ? WHERE id = ?', (recipe_id, existing['id']))
    else:
        base = conn.execute(
            '''SELECT id FROM meal_plan_days
               WHERE meal_plan_id = ? AND week_number = ? AND day_of_week = ? AND meal_type = 'lounas' ''',
            (plan_id, week, day_of_week)).fetchone()
        if not base:
            conn.close()
            return jsonify({'error': 'Päivää ei löytynyt tältä ruokalistalta'}), 404
        conn.execute(
            f'''INSERT INTO meal_plan_days (meal_plan_id, week_number, day_of_week, recipe_id, meal_type)
                VALUES (?, ?, ?, ?, '{meal_type}')''',
            (plan_id, week, day_of_week, recipe_id))
    conn.commit()
    conn.close()
    return jsonify({'message': f'{label} tallennettu'})


@app.route('/api/plans/<int:plan_id>/second-main', methods=['POST'])
def update_second_main(plan_id):
    """Valinnainen toinen pääruoka yhdelle päivälle (esim. Kymenkartanon
    ruokalistalla nähty kahden vaihtoehdon päivä)."""
    return _update_extra_main(plan_id, 'lounas2', 'Toinen pääruoka')


@app.route('/api/plans/<int:plan_id>/third-main', methods=['POST'])
def update_third_main(plan_id):
    """Valinnainen kolmas pääruoka — käytännössä vain Kestin perjantaille:
    Kesti tarjoaa perjantaisin aina 2 pääruokaa (lounas + lounas2, molemmat
    näkyvät Kestin omassa menussa), ja tämä kolmas rivi on Hoivaa varten —
    Hoivan ma-pe peilaa Kestin dataa sellaisenaan, joten tallentamalla tämä
    Kestin perjantairiville Hoiva saa sen automaattisesti mukaan omaan
    menuunsa ilman että Hoiva tarvitsee mitään omaa dataa. Kestin omat
    vientipohjat eivät koskaan näytä tätä riviä, vain Hoivan pohjat."""
    return _update_extra_main(plan_id, 'lounas3', 'Kolmas pääruoka')


@app.route('/api/plans/<int:plan_id>/sauce', methods=['POST'])
def update_sauce(plan_id):
    """Valinnainen kastike-lisäys yhdelle päivälle — koskee vain Kestiä."""
    return _update_extra_main(plan_id, 'kastike', 'Kastike', require_kesti=True)


@app.route('/api/plans/<int:plan_id>/veg-side', methods=['POST'])
def update_veg_side(plan_id):
    """Valinnainen kasvislisäke yhdelle päivälle — koskee vain Kestiä."""
    return _update_extra_main(plan_id, 'kasvislisäke', 'Kasvislisäke', require_kesti=True)


@app.route('/api/plans/<int:plan_id>/energy-side', methods=['POST'])
def update_energy_side(plan_id):
    """Valinnainen energialisäke yhdelle päivälle — koskee vain Kestiä."""
    return _update_extra_main(plan_id, 'energialisäke', 'Energialisäke', require_kesti=True)


@app.route('/api/recipes/random')
def random_recipe():
    """Return one random recipe of the given recipe_type, for the
    drag-and-drop accompaniment tray's suggestion chips."""
    recipe_type = request.args.get('recipe_type', '')
    if recipe_type not in RECIPE_TYPES:
        return jsonify({'error': 'Tuntematon reseptityyppi'}), 400
    conn = get_db()
    row = conn.execute(
        'SELECT id, name_fi FROM recipes WHERE recipe_type = ? ORDER BY RANDOM() LIMIT 1',
        (recipe_type,)).fetchone()
    conn.close()
    if not row:
        return jsonify({'error': 'Reseptejä ei löytynyt'}), 404
    return jsonify({'id': row['id'], 'name_fi': row['name_fi']})


@app.route('/api/plans/<int:plan_id>/suggestions')
def meal_suggestions(plan_id):
    """Suggest same-category replacements for a slot."""
    try:
        week = _int_arg(request.args, 'week')
        slot = _int_arg(request.args, 'slot')
    except _BadIntArg as e:
        return jsonify({'error': e.message}), 400
    meal_type = request.args.get('meal_type', 'lounas')
    modifier = MealModifier(DB_PATH)
    suggestions = modifier.suggest_recipe_replacement(plan_id, week, slot, meal_type)
    return jsonify([{'id': s[0], 'name': s[1]} for s in suggestions])


@app.route('/api/plans/<int:plan_id>/history')
def plan_history(plan_id):
    modifier = MealModifier(DB_PATH)
    return jsonify(modifier.get_modification_history(plan_id))


@app.route('/api/plans/<int:plan_id>/export')
def export_plan(plan_id):
    """Export meal plan to Excel in the wall-chart format."""
    out = os.path.join(OUTPUT_DIR, f'ruokalista_{plan_id}.xlsx')
    exporter = MealPlanExporter(DB_PATH)
    path = exporter.export_meal_plan(plan_id, out)
    if not path:
        return jsonify({'error': 'Vienti epäonnistui'}), 500
    return send_file(path, as_attachment=True,
                     download_name=f'ruokalista_{plan_id}.xlsx')



@app.route('/api/plans/<int:plan_id>/menu')
def export_menu(plan_id):
    """Generate a weekly lunch menu in the restaurant's branded template.

    Default: branded PDF (reuses the restaurant's own menu graphics).
    ?format=docx: plain editable Word version (no graphics) as a fallback.
    """
    from datetime import datetime as _dt
    fmt = request.args.get('format', 'pdf')
    try:
        week = _int_arg(request.args, 'week', 1)
        year = _int_arg(request.args, 'year', _dt.now().year)
    except _BadIntArg as e:
        return jsonify({'error': e.message}), 400

    conn = get_db()
    plan = conn.execute('SELECT name, facility FROM meal_plans WHERE id = ?', (plan_id,)).fetchone()
    if not plan:
        conn.close()
        return jsonify({'error': 'Ruokalistaa ei löydy'}), 404
    facility = plan['facility'] or 'kesti'
    days_per_week = FACILITY_DAYS_PER_WEEK.get(facility, 5)
    # Käsin tyhjennetyt paikat (ks. /clear_meal) jätetään pois viennistä —
    # niillä ei ole reseptiä johon viitata.
    day_rows = [r for r in _effective_meal_plan_days(conn, plan_id, week=week) if r['recipe_id'] is not None]
    recipe_ids = list({r['recipe_id'] for r in day_rows})
    recipes_by_id = {}
    if recipe_ids:
        placeholders = ','.join('?' * len(recipe_ids))
        recipes_by_id = {rr['id']: rr for rr in conn.execute(
            f'SELECT id, name_fi, dish_category, notes FROM recipes WHERE id IN ({placeholders})',
            recipe_ids).fetchall()}
    conn.close()
    if not day_rows:
        return jsonify({'error': f'Viikolle {week} ei löytynyt aterioita'}), 404
    rows = [{'day_of_week': r['day_of_week'], 'meal_type': r['meal_type'], 'dessert_text': r['dessert_text'],
            'name_fi': recipes_by_id[r['recipe_id']]['name_fi'],
            'dish_category': recipes_by_id[r['recipe_id']]['dish_category'],
            'notes': recipes_by_id[r['recipe_id']]['notes']} for r in day_rows]

    # Real per-weekday soup/main/salad, built directly from the generator's
    # own assignment — no name-pattern heuristics needed, the roles are
    # already known (meal_type: 'keitto'/'salaatti'/'lounas').
    meals = [{'soup': None, 'main': None, 'main2': None, 'main3': None, 'salad': None, 'sides': None,
             'sauce': None, 'veg_side': None, 'energy_side': None} for _ in range(days_per_week)]
    dessert_text = None
    for r in rows:
        weekday = r['day_of_week']
        if not (0 <= weekday < days_per_week):
            continue
        role = MEAL_ROLE_BY_TYPE.get(r['meal_type'], 'main')
        meals[weekday][role] = {'name': r['name_fi'], 'category': r['dish_category'],
                                'notes': r['notes']}
        if weekday == DESSERT_DAY_BY_FACILITY.get(facility) and r['dessert_text']:
            dessert_text = r['dessert_text']

    if facility == 'kesti':
        # Kesti's own menu never shows the 3rd main at all (it exists only
        # so Hoiva's mirrored Friday can show it) — and shows the 2nd main
        # only on Friday specifically, since Kesti always serves 2 mains
        # that day; a main2 manually added to any other day stays hidden
        # from Kesti's own menu (Hoiva still sees it via the live mirror).
        for weekday, day_meals in enumerate(meals):
            day_meals['main3'] = None
            if weekday != 4:
                day_meals['main2'] = None
            extra_names = [day_meals[role]['name'] for role in ACCOMPANIMENT_ROLES if day_meals.get(role)]
            day_meals['sides'] = {'name': f"LISÄKKEET: {', '.join(extra_names)}", 'notes': None} if extra_names else None

    if fmt == 'docx':
        try:
            from menu_generator import build_week_menu
        except ImportError:
            return jsonify({'error': 'python-docx-kirjasto puuttuu. Aja KAYNNISTA.bat uudelleen.'}), 500
        out = os.path.join(OUTPUT_DIR, f'viikkomenu_plan{plan_id}_vko{week}.docx')
        # main2/main3 visibility is already decided above (per facility/day)
        # by what's actually populated in `meals` — build_week_menu just
        # renders whatever it's given.
        path, _ = build_week_menu(meals, week, year, out, day_names=WEEKDAY_LABELS[:days_per_week])
        download_name = f'viikkomenu_vko{week}.docx'
    elif facility == 'hoiva':
        variant = request.args.get('variant', 'tammikoti')
        if variant not in ('tammikoti', 'asukas'):
            return jsonify({'error': 'Tuntematon pohja'}), 400
        try:
            from menu_pdf_generator import build_hoiva_menu_pdf
        except ImportError as e:
            return jsonify({'error': f'PDF-kirjasto puuttuu ({e}). Aja KAYNNISTA.bat uudelleen.'}), 500
        out = os.path.join(OUTPUT_DIR, f'viikkomenu_{variant}_plan{plan_id}_vko{week}.pdf')
        build_hoiva_menu_pdf(meals, week, year, out, variant=variant,
                             dessert_day=DESSERT_DAY_BY_FACILITY.get('hoiva'),
                             dessert_text=dessert_text)
        path = out
        download_name = f'viikkomenu_{variant}_vko{week}.pdf'
    elif facility == 'kymenkartano':
        try:
            from menu_pdf_generator import build_kymenkartano_menu_pdf
        except ImportError as e:
            return jsonify({'error': f'PDF-kirjasto puuttuu ({e}). Aja KAYNNISTA.bat uudelleen.'}), 500
        out = os.path.join(OUTPUT_DIR, f'viikkomenu_plan{plan_id}_vko{week}.pdf')
        build_kymenkartano_menu_pdf(meals, week, year, out,
                                    dessert_day=DESSERT_DAY_BY_FACILITY.get('kymenkartano'),
                                    dessert_text=dessert_text)
        path = out
        download_name = f'viikkomenu_vko{week}.pdf'
    else:
        if facility != 'kesti':
            return jsonify({'error': f'{FACILITIES.get(facility, facility)}-toimipisteelle ei ole vielä '
                                     'omaa graafista PDF-pohjaa — käytä toistaiseksi Word-vientiä.'}), 400
        try:
            from menu_pdf_generator import build_week_menu_pdf
        except ImportError as e:
            return jsonify({'error': f'PDF-kirjasto puuttuu ({e}). Aja KAYNNISTA.bat uudelleen.'}), 500
        out = os.path.join(OUTPUT_DIR, f'viikkomenu_plan{plan_id}_vko{week}.pdf')
        path, _ = build_week_menu_pdf(meals, week, year, out)
        download_name = f'viikkomenu_vko{week}.pdf'

    # Työpöytäversion sisäinen ikkuna (pywebview/WebView2) ei tue selaimen
    # normaalia lataustoimintoa — JS:llä laukaistu blob-lataus ei näy
    # käyttäjälle mitenkään, vaikka tiedosto syntyy palvelimella onnistuneesti.
    # Palvelin ja käyttöliittymä ajavat samalla koneella, joten kopioidaan
    # tiedosto suoraan käyttäjän oikeaan Lataukset-kansioon sen sijaan että
    # yritettäisiin selaimen kautta tapahtuvaa latausta.
    saved_path = _save_to_downloads(path, download_name)

    return jsonify({
        'message': f'Viikkomenu tallennettu Lataukset-kansioon: {download_name}',
        'filename': download_name,
        'saved_path': saved_path,
    })


@app.route('/api/plans/<int:plan_id>/instructions-pdf')
def export_instructions_pdf(plan_id):
    """Print all cooking instructions for one week's meals, grouped by
    day, for kitchen staff — uses the real per-recipe instructions column
    (not the ticket's assumed 'notes' field, which holds diet/allergen
    codes, and not the dead recipes.ingredients JSON)."""
    try:
        week = _int_arg(request.args, 'week', 1)
        target_servings = _optional_int_arg(request.args, 'servings')
    except _BadIntArg as e:
        return jsonify({'error': e.message}), 400
    if target_servings is not None and not (1 <= target_servings <= 2000):
        return jsonify({'error': 'Annosmäärän pitää olla 1-2000'}), 400

    conn = get_db()
    plan = conn.execute('SELECT name, facility FROM meal_plans WHERE id = ?', (plan_id,)).fetchone()
    if not plan:
        conn.close()
        return jsonify({'error': 'Ruokalistaa ei löydy'}), 404
    _ensure_ingredient_editor_tables(conn)  # varmistaa recipe_ingredients.sort_order-sarakkeen olemassaolon
    facility = plan['facility'] or 'kesti'
    days_per_week = FACILITY_DAYS_PER_WEEK.get(facility, 5)
    # Käsin tyhjennetyt paikat (ks. /clear_meal) jätetään pois viennistä —
    # niillä ei ole reseptiä johon viitata.
    day_rows = [r for r in _effective_meal_plan_days(conn, plan_id, week=week) if r['recipe_id'] is not None]
    recipe_ids = list({r['recipe_id'] for r in day_rows})
    recipes_by_id = {}
    ingredients_by_recipe = defaultdict(list)
    if recipe_ids:
        placeholders = ','.join('?' * len(recipe_ids))
        recipes_by_id = {rr['id']: rr for rr in conn.execute(
            f'SELECT id, name_fi, instructions, servings FROM recipes WHERE id IN ({placeholders})',
            recipe_ids).fetchall()}
        for rr in conn.execute(
                f'''SELECT ri.recipe_id, i.name_fi, ri.quantity, ri.unit
                    FROM recipe_ingredients ri JOIN ingredients i ON i.id = ri.ingredient_id
                    WHERE ri.recipe_id IN ({placeholders})
                    ORDER BY ri.sort_order, i.name_fi''', recipe_ids).fetchall():
            ingredients_by_recipe[rr['recipe_id']].append(rr)
    conn.close()
    if not day_rows:
        return jsonify({'error': f'Viikolle {week} ei löytynyt aterioita'}), 404

    role_labels = {'keitto': 'Keitto', 'salaatti': 'Salaatti'}
    days_data = [[] for _ in range(days_per_week)]
    missing_servings = set()
    for r in day_rows:
        weekday = r['day_of_week']
        if not (0 <= weekday < days_per_week):
            continue
        rr = recipes_by_id[r['recipe_id']]
        # Ei oletusarvoa puuttuvalle annosmäärälle: monet reseptit ilman
        # annosmäärää osoittautuivat jo valmiiksi tuotantoerän kokoisiksi
        # (kg-mittakaavan raaka-aineita), joten oletus "4 annosta" ja siitä
        # ylöspäin skaalaus tuotti järjettömän suuria tilausmääriä käytännössä.
        if target_servings and rr['servings']:
            factor = target_servings / rr['servings']
        else:
            factor = 1
            if target_servings and not rr['servings']:
                missing_servings.add(rr['name_fi'])
        ingredients = [{'name_fi': ing['name_fi'],
                        'quantity': round(ing['quantity'] * factor, 3) if ing['quantity'] is not None else None,
                        'unit': ing['unit']}
                      for ing in ingredients_by_recipe.get(r['recipe_id'], [])]
        days_data[weekday].append({
            'role_label': role_labels.get(r['meal_type'], 'Pääruoka'),
            'name_fi': rr['name_fi'],
            'instructions': rr['instructions'],
            'ingredients': ingredients,
        })

    try:
        from menu_pdf_generator import build_kitchen_instructions_pdf
    except ImportError as e:
        return jsonify({'error': f'PDF-kirjasto puuttuu ({e}). Aja KAYNNISTA.bat uudelleen.'}), 500

    out = os.path.join(OUTPUT_DIR, f'valmistusohjeet_plan{plan_id}_vko{week}.pdf')
    build_kitchen_instructions_pdf(week, WEEKDAY_LABELS[:days_per_week], days_data, out)
    download_name = f'valmistusohjeet_vko{week}.pdf'
    saved_path = _save_to_downloads(out, download_name)

    message = f'Valmistusohjeet tallennettu Lataukset-kansioon: {download_name}'
    if missing_servings:
        message += (f' — huom: {len(missing_servings)} reseptillä ei annosmäärää, '
                    f'niiden raaka-aineita EI skaalattu: ' +
                    ', '.join(sorted(missing_servings)[:8]) +
                    ('…' if len(missing_servings) > 8 else ''))
    return jsonify({
        'message': message,
        'filename': download_name,
        'saved_path': saved_path,
    })


@app.route('/api/open-downloads-folder', methods=['POST'])
def open_downloads_folder():
    """Avaa käyttäjän Lataukset-kansion Resurssienhallinnassa — kätevä
    jatko sille, että viedyt tiedostot kirjoitetaan suoraan sinne (ks.
    _save_to_downloads), koska pywebviewin sisäinen ikkuna ei tarjoa mitään
    muuta tapaa nähdä missä ladattu tiedosto on."""
    try:
        os.startfile(_downloads_dir())
        return jsonify({'message': 'Lataukset-kansio avattu'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/plans/<int:plan_id>/kespro')
def export_kespro(plan_id):
    """Generate a weekly Kespro order template (Excel) for a given week."""
    try:
        week = _int_arg(request.args, 'week', 1)
        target_servings = _optional_int_arg(request.args, 'servings')
    except _BadIntArg as e:
        return jsonify({'error': e.message}), 400
    if target_servings is not None and not (1 <= target_servings <= 2000):
        return jsonify({'error': 'Annosmäärän pitää olla 1-2000'}), 400
    from openpyxl import Workbook
    from openpyxl.styles import Font

    conn = get_db()
    plan = conn.execute('SELECT name FROM meal_plans WHERE id = ?', (plan_id,)).fetchone()
    if not plan:
        conn.close()
        return jsonify({'error': 'Ruokalistaa ei löydy'}), 404

    # Mirror-aware: for Hoiva this must include the linked Kesti plan's
    # ma-pe recipes, not just Hoiva's own (la-su-only) rows. recipe_id_list
    # keeps duplicates (same recipe used on 2 different days that week
    # must be counted/ordered twice), distinct_ids is only for the lookups.
    # Käsin tyhjennetyt paikat (ks. /clear_meal) jätetään pois tilauslistasta —
    # niillä ei ole reseptiä johon viitata.
    day_rows = [r for r in _effective_meal_plan_days(conn, plan_id, week=week) if r['recipe_id'] is not None]
    recipe_id_list = [r['recipe_id'] for r in day_rows]
    distinct_ids = list(set(recipe_id_list))

    recipe_info = {}
    ing_by_recipe = defaultdict(list)
    if distinct_ids:
        placeholders = ','.join('?' * len(distinct_ids))
        recipe_info = {rr['id']: rr for rr in conn.execute(
            f'SELECT id, name_fi, dish_category, servings FROM recipes WHERE id IN ({placeholders})',
            distinct_ids).fetchall()}
        for rr in conn.execute(
                f'''SELECT ri.recipe_id, i.name_fi, ri.quantity, ri.unit
                    FROM recipe_ingredients ri JOIN ingredients i ON i.id = ri.ingredient_id
                    WHERE ri.recipe_id IN ({placeholders})''', distinct_ids).fetchall():
            ing_by_recipe[rr['recipe_id']].append(rr)
    conn.close()

    rows = [{'day_of_week': r['day_of_week'], 'name_fi': recipe_info[r['recipe_id']]['name_fi'],
            'dish_category': recipe_info[r['recipe_id']]['dish_category']} for r in day_rows]

    # Aggregated ingredient quantities (when recipes have linked ingredients),
    # scaled per recipe to target_servings if given — recipes are written
    # for whatever batch they were originally portioned for (commonly
    # 4-10 servings), not the restaurant's actual daily headcount. Recipes
    # with no servings value are left unscaled (factor 1): many turned out
    # to already be bulk/production-batch sized (kg-scale ingredients), and
    # assuming a small default like 4 produced wildly inflated orders.
    missing_servings = set()
    agg = defaultdict(lambda: {'total': 0.0, 'unit': '', 'has_qty': False})
    for rid in recipe_id_list:
        if target_servings and recipe_info[rid]['servings']:
            factor = target_servings / recipe_info[rid]['servings']
        else:
            factor = 1
            if target_servings and not recipe_info[rid]['servings']:
                missing_servings.add(recipe_info[rid]['name_fi'])
        for ing in ing_by_recipe.get(rid, []):
            key = (ing['name_fi'], ing['unit'])
            if ing['quantity'] is not None:
                agg[key]['total'] += ing['quantity'] * factor
                agg[key]['has_qty'] = True
            agg[key]['unit'] = ing['unit']
    ing_rows = [{'name_fi': k[0], 'unit': k[1], 'total': round(v['total'], 3) if v['has_qty'] else None}
               for k, v in sorted(agg.items())]

    wb = Workbook()
    ws = wb.active
    ws.title = f'Viikko {week}'
    ws['A1'] = f'KESPRO-TILAUSPOHJA — {plan["name"]} — Viikko {week}'
    ws['A1'].font = Font(bold=True, size=13)
    servings_note = f' — mitoitettu {target_servings} annokselle' if target_servings else ''
    ws['A2'] = f'Luotu: {datetime.now().strftime("%d.%m.%Y %H:%M")}{servings_note}'
    if missing_servings:
        ws['A3'] = (f'Huom: {len(missing_servings)} reseptillä ei annosmäärää — '
                    f'niiden raaka-aineita EI skaalattu: ' +
                    ', '.join(sorted(missing_servings)[:8]) +
                    ('…' if len(missing_servings) > 8 else ''))

    headers = ['Kespro-tuotenumero', 'Tuote', 'Yksikkö', 'Määrä', 'Toimituspäivä', 'Huomiot']
    for col, h in enumerate(headers, 1):
        c = ws.cell(4, col, h)
        c.font = Font(bold=True)

    row = 5
    if ing_rows:
        delivery = (datetime.now() + timedelta(days=7)).strftime('%d.%m.%Y')
        for ing in ing_rows:
            ws.cell(row, 1, '')  # SKU filled in by staff / future mapping
            ws.cell(row, 2, ing['name_fi'])
            ws.cell(row, 3, ing['unit'])
            ws.cell(row, 4, round(ing['total'], 2) if ing['total'] is not None else '')
            ws.cell(row, 5, delivery)
            row += 1
    else:
        ws.cell(row, 2, 'Reseptien raaka-aineita ei ole vielä syötetty järjestelmään.')
        row += 2
        ws.cell(row, 2, 'Viikon ruoat (käytä tilauksen pohjana):')
        ws.cell(row, 2).font = Font(bold=True)
        row += 1
        for r in rows:
            ws.cell(row, 2, r['name_fi'])
            ws.cell(row, 3, CATEGORIES.get(r['dish_category'], ''))
            row += 1

    for col, width in zip('ABCDEF', (20, 40, 10, 10, 15, 30)):
        ws.column_dimensions[col].width = width

    out = os.path.join(OUTPUT_DIR, f'kespro_tilaus_plan{plan_id}_vko{week}.xlsx')
    wb.save(out)
    return send_file(out, as_attachment=True,
                     download_name=f'kespro_tilaus_vko{week}.xlsx')


# ============================================================
# IMPORT: PoweResta Excel + pasted text (no scraping needed)
# ============================================================

def _parse_qty(v):
    """PoweRestan painosarake: 2.5, '0,500', '0,226 kg', '0,906 l' -> (määrä, yksikkö) tai (None, None)."""
    import re as _re
    if v is None or v == '':
        return None, None
    s = str(v).strip().replace(',', '.')
    m = _re.match(r'^([\d. ]+?)\s*([a-zA-Z]*)$', s)
    if not m:
        return None, None
    try:
        qty = round(float(m.group(1).replace(' ', '')), 3)
    except ValueError:
        return None, None
    return qty, (m.group(2) or 'kg').lower()


def _save_recipe_ingredients(recipe_id, items):
    """Tallenna reseptin raaka-aineet määrineen (ohittaa jos jo tallennettu)."""
    if not recipe_id or not items:
        return
    conn = get_db()
    try:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS ingredients (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name_fi TEXT UNIQUE NOT NULL);
            CREATE TABLE IF NOT EXISTS recipe_ingredients (
                recipe_id INTEGER NOT NULL,
                ingredient_id INTEGER NOT NULL,
                quantity REAL, unit TEXT,
                UNIQUE (recipe_id, ingredient_id));
        """)
        has = conn.execute('SELECT COUNT(*) FROM recipe_ingredients WHERE recipe_id=?',
                           (recipe_id,)).fetchone()[0]
        if has:
            return  # älä tuplaa
        for it in items:
            conn.execute('INSERT OR IGNORE INTO ingredients (name_fi) VALUES (?)',
                         (it['name'],))
            iid = conn.execute('SELECT id FROM ingredients WHERE name_fi=?',
                               (it['name'],)).fetchone()[0]
            conn.execute('''INSERT OR IGNORE INTO recipe_ingredients
                            (recipe_id, ingredient_id, quantity, unit)
                            VALUES (?,?,?,?)''',
                         (recipe_id, iid, it['qty'], it.get('unit', 'kg')))
        conn.commit()
    finally:
        conn.close()


def _replace_recipe_ingredients(recipe_id, items):
    """Replace a recipe's ingredients and record the imported differences."""
    if not recipe_id:
        return
    conn = get_db()
    try:
        _ensure_ingredient_editor_tables(conn)
        old_rows = conn.execute(
            '''SELECT i.name_fi, ri.quantity, ri.unit
               FROM recipe_ingredients ri JOIN ingredients i ON i.id=ri.ingredient_id
               WHERE ri.recipe_id=?''', (recipe_id,)).fetchall()
        old = {row['name_fi'].strip().lower(): row for row in old_rows}
        new = {str(item.get('name', '')).strip().lower(): item for item in items
               if str(item.get('name', '')).strip()}

        for key, row in old.items():
            item = new.get(key)
            if item is None:
                _log_ingredient_change(conn, recipe_id, 'removed', row['name_fi'],
                                       old_quantity=row['quantity'], old_unit=row['unit'])
            elif row['quantity'] != item.get('qty') or (row['unit'] or '') != (item.get('unit') or 'kg'):
                _log_ingredient_change(conn, recipe_id, 'updated', row['name_fi'],
                                       old_quantity=row['quantity'], new_quantity=item.get('qty'),
                                       old_unit=row['unit'], new_unit=item.get('unit', 'kg'))
        for key, item in new.items():
            if key not in old:
                _log_ingredient_change(conn, recipe_id, 'added', item['name'],
                                       new_quantity=item.get('qty'), new_unit=item.get('unit', 'kg'))

        conn.execute('DELETE FROM recipe_ingredients WHERE recipe_id=?', (recipe_id,))
        for item in new.values():
            iid = _get_or_create_ingredient_id(conn, item['name'])
            conn.execute('''INSERT INTO recipe_ingredients (recipe_id, ingredient_id, quantity, unit)
                            VALUES (?,?,?,?)''',
                         (recipe_id, iid, item.get('qty'), item.get('unit', 'kg')))
        conn.commit()
    finally:
        conn.close()



def _queue_append(items):
    """Append recipes to the review queue, dedup by name."""
    try:
        with open(REVIEW_FILE, encoding='utf-8') as f:
            queue = json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        queue = []
    existing_names = {(r.get('name_fi') or '').strip().lower() for r in queue}
    added = 0
    for item in items:
        key = (item.get('name_fi') or '').strip().lower()
        if key and key not in existing_names:
            queue.append(item)
            existing_names.add(key)
            added += 1
    with open(REVIEW_FILE, 'w', encoding='utf-8') as f:
        json.dump(queue, f, ensure_ascii=False, indent=2)
    return added


def _extract_poweresta_pdf_recipes_legacy(pdf_bytes, filename):
    """Extract text-based PoweResta recipe PDFs into review-queue items.

    PoweResta layouts vary by version, so labels (Nimi/Ruokavaliot/Rivit) are
    used when present and otherwise the first meaningful line becomes the name.
    Scanned PDFs are rejected clearly because they need OCR before import.
    """
    import re as _re
    import uuid
    import fitz
    from recipe_scraper_v2 import guess_category, guess_season

    try:
        document = fitz.open(stream=pdf_bytes, filetype='pdf')
        text = '\n'.join(page.get_text('text') for page in document)
        document.close()
    except Exception as exc:
        raise ValueError(f'PDF:n avaus epäonnistui: {exc}') from exc

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        raise ValueError('PDF:stä ei löytynyt tekstisisältöä. Skannattu PDF pitää OCR-käsitellä ensin.')

    name = ''
    diet = ''
    for index, line in enumerate(lines):
        match = _re.match(r'^Nimi\s*[:：-]?\s*(.+)$', line, _re.IGNORECASE)
        if match:
            name = match.group(1).strip()
            break
        if line.lower() == 'nimi' and index + 1 < len(lines):
            name = lines[index + 1]
            break
    if not name:
        ignored = {'resepti', 'reseptikortti', 'raaka-aineet', 'ainekset', 'ohje', 'valmistusohje'}
        name = next((line for line in lines if line.lower() not in ignored), '')
    if not name or len(name) > 160:
        raise ValueError('Reseptin nimeä ei voitu tunnistaa PDF:stä')

    for index, line in enumerate(lines):
        match = _re.match(r'^Ruokavaliot\s*[:：-]?\s*(.+)$', line, _re.IGNORECASE)
        if match:
            diet = match.group(1).strip()
            break
        if line.lower() == 'ruokavaliot' and index + 1 < len(lines):
            diet = lines[index + 1]
            break

    ingredients, instructions = [], []
    qty_pattern = _re.compile(r'^(?:[•*-]\s*)?(\d+(?:[.,]\d+)?)\s*(kg|g|l|dl|ml|cl|kpl|rkl|tl)\s+(.+)$', _re.IGNORECASE)
    in_ingredients = False
    ingredient_headers = ('rivit', 'raaka-aineet', 'ainekset', 'ostopaino')
    instruction_headers = ('valmistusohje', 'ohje', 'valmistus')
    for line in lines:
        lower = line.lower().rstrip(':')
        if any(header in lower for header in instruction_headers):
            in_ingredients = False
            continue
        if any(header in lower for header in ingredient_headers):
            in_ingredients = True
            continue
        match = qty_pattern.match(line)
        if match:
            in_ingredients = True
            qty, unit = _parse_qty(f'{match.group(1)} {match.group(2)}')
            ingredients.append({'name': match.group(3).strip(), 'qty': qty, 'unit': unit})
        elif in_ingredients and not ingredients:
            continue  # table headings before the first ingredient
        elif not in_ingredients and line != name:
            instructions.append(line)

    if not ingredients:
        raise ValueError('PDF:stä ei löytynyt raaka-ainerivejä määrineen')
    ing_text = ' '.join(item['name'] for item in ingredients)
    return [{
        'queue_id': str(uuid.uuid4()),
        'name_fi': name,
        'source_url': '',
        'source_site': f'PoweResta PDF ({filename})',
        'ingredients_raw': [item['name'] for item in ingredients],
        'ingredients_struct': ingredients,
        'instructions_raw': '\n'.join(instructions),
        'season': guess_season(name, ing_text) or 'kaikki',
        'meal_type': 'lounas',
        'dish_category': guess_category(name, ing_text),
        'prep_time_min': None,
        'difficulty': None,
        'servings': None,
        'notes': f'Ruokavaliot: {diet}' if diet else '',
        'scraped_at': datetime.now().isoformat(),
        'review_action': 'created',
    }]


def _poweresta_pdf_item(name, ingredients, instructions, filename):
    """Build the common review-queue representation for a parsed PDF recipe."""
    from recipe_scraper_v2 import guess_category, guess_season
    import uuid
    ing_text = ' '.join(item['name'] for item in ingredients)
    return {
        'queue_id': str(uuid.uuid4()), 'name_fi': name,
        'source_url': '', 'source_site': f'PoweResta PDF ({filename})',
        'ingredients_raw': [item['name'] for item in ingredients],
        'ingredients_struct': ingredients, 'instructions_raw': '\n'.join(instructions),
        'season': guess_season(name, ing_text) or 'kaikki', 'meal_type': 'lounas',
        'dish_category': guess_category(name, ing_text), 'prep_time_min': None,
        'difficulty': None, 'servings': None, 'notes': '',
        'scraped_at': datetime.now().isoformat(), 'review_action': 'created',
    }


def _extract_poweresta_pdf_recipes(pdf_bytes, filename):
    """Parse the two PoweResta PDF layouts used by this organisation.

    Newer exports put ``0,325 kg`` and the ingredient on consecutive lines,
    one recipe per page. Older exports put purchase/use weights and units on
    separate rows before a possibly multi-line ingredient name.
    """
    import re as _re
    import fitz
    try:
        document = fitz.open(stream=pdf_bytes, filetype='pdf')
    except Exception as exc:
        raise ValueError(f'PDF:n avaus epäonnistui: {exc}') from exc

    recipes = []
    quantity_unit = _re.compile(r'^(\d+(?:[.,]\d+)?)\s*(kg|g|l|dl|ml|cl|kpl)$', _re.I)
    quantity_only = _re.compile(r'^\d+(?:[.,]\d+)?$')
    for page in document:
        lines = [line.strip() for line in page.get_text('text').splitlines() if line.strip()]
        if not lines:
            continue
        joined = '\n'.join(lines).lower()
        instructions, ingredients, name = [], [], ''

        # New PoweResta layout: ingredient amount and name are adjacent.
        if 'raaka-aineet yht:' in joined and 'käyttöpaino' in joined:
            start = next((i for i, line in enumerate(lines) if line.lower() == 'käyttöpaino'), -1)
            end = next((i for i, line in enumerate(lines) if line.lower().startswith('tuotepakkaukset')), len(lines))
            for i in range(start + 1, end - 1):
                match = quantity_unit.match(lines[i])
                if not match:
                    continue
                candidate = lines[i + 1]
                # Stage yields and table values are followed by another number,
                # whereas real ingredient rows are followed by a name.
                if quantity_unit.match(candidate) or quantity_only.match(candidate):
                    continue
                if candidate.lower().startswith(('tuotepakkaukset', 'saanto', 'raakapaino')):
                    continue
                qty, unit = _parse_qty(f'{match.group(1)} {match.group(2)}')
                ingredients.append({'name': candidate, 'qty': qty, 'unit': unit})
            name_index = next((i for i, line in enumerate(lines) if line.lower().startswith('allergeenit:')), -1)
            if name_index >= 0:
                name = next((line for line in lines[name_index + 1:]
                             if not line.lower().startswith(('rakenteet:', 'hallinto'))), '')
            instructions = lines[end:name_index if name_index >= 0 else len(lines)]

        # Older PoweResta layout: purchase/use weight, unit, then ingredient.
        elif 'reseptin työohje' in joined and 'ostopaino' in joined:
            start = next((i for i, line in enumerate(lines) if line.lower() == 'ostopaino'), -1)
            end = next((i for i, line in enumerate(lines) if line.lower().startswith('raakapaino yht')), len(lines))
            i = start + 1
            while i + 4 < end:
                if (quantity_only.match(lines[i]) and quantity_only.match(lines[i + 1])
                        and lines[i + 2].lower() in {'kg', 'g', 'l', 'dl', 'ml', 'kpl'}
                        and lines[i + 3].lower() == lines[i + 2].lower()):
                    qty, unit = _parse_qty(f'{lines[i]} {lines[i + 2]}')
                    j = i + 4
                    name_lines = []
                    while j < end:
                        if (j + 3 < end and quantity_only.match(lines[j]) and quantity_only.match(lines[j + 1])
                                and lines[j + 2].lower() in {'kg', 'g', 'l', 'dl', 'ml', 'kpl'}
                                and lines[j + 3].lower() == lines[j + 2].lower()):
                            break
                        if lines[j].upper() in {'PATA', 'KYLMÄSÄILYTYS', 'KYPSENNYS'}:
                            break
                        name_lines.append(lines[j])
                        j += 1
                    ingredient_name = ' '.join(name_lines).strip()
                    if ingredient_name and 'vuoka' not in ingredient_name.lower():
                        ingredients.append({'name': ingredient_name, 'qty': qty, 'unit': unit})
                    i = max(j, i + 4)
                else:
                    i += 1
            serving_index = next((i for i, line in enumerate(lines) if _re.match(r'^\d+ annosta', line, _re.I)), -1)
            if serving_index > 0:
                name = lines[serving_index - 1]

        if name and ingredients:
            recipes.append(_poweresta_pdf_item(name, ingredients, instructions, filename))
    document.close()
    if not recipes:
        raise ValueError('PDF:stä ei löytynyt tunnistettavaa PoweResta-reseptiä tai raaka-aineita')
    return recipes


@app.route('/api/import/poweresta-pdf', methods=['POST'])
def import_poweresta_pdf():
    """Import one text-based PoweResta recipe PDF into the review queue."""
    if 'file' not in request.files or not request.files['file'].filename:
        return jsonify({'error': 'Valitse PDF-tiedosto'}), 400
    uploaded = request.files['file']
    if not uploaded.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Vain PDF-tiedostot ovat sallittuja'}), 400
    try:
        recipes = _extract_poweresta_pdf_recipes(uploaded.read(), uploaded.filename)
    except ValueError as exc:
        return jsonify({'error': str(exc)}), 400
    except Exception:
        app.logger.exception('PoweResta PDF import failed')
        return jsonify({'error': 'PDF:n käsittely epäonnistui'}), 500

    conn = get_db()
    try:
        for recipe in recipes:
            existing = conn.execute('SELECT id FROM recipes WHERE name_fi=?',
                                    (recipe['name_fi'],)).fetchone()
            if existing:
                recipe['review_action'] = 'updated'
                recipe['existing_recipe_id'] = existing['id']
    finally:
        conn.close()
    added = _queue_append(recipes)
    action = recipes[0]['review_action']
    return jsonify({'success': True, 'recipe_name': recipes[0]['name_fi'],
                    'action': action, 'queue_id': recipes[0]['queue_id'],
                    'total_processed': len(recipes), 'added': added})


@app.route('/api/import/poweresta', methods=['POST'])
def import_poweresta():
    """
    Import recipes from a PoweResta-format Excel file
    (the output of the recipe PDF converter project).

    Format per sheet: 'Nimi' label with recipe name beside it,
    'Ruokavaliot' with diet codes, ingredient names in column C
    below the 'Rivit' header row.
    """
    from openpyxl import load_workbook
    from recipe_scraper_v2 import guess_category, guess_season
    from datetime import datetime as _dt

    if 'file' not in request.files or not request.files['file'].filename:
        return jsonify({'error': 'Valitse Excel-tiedosto'}), 400
    f = request.files['file']
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({'error': 'Tiedoston pitää olla .xlsx'}), 400

    import tempfile
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    tmp.close()  # Windows lukitsee avoimet tiedostot — vapauta kahva heti
    f.save(tmp.name)

    try:
        wb = load_workbook(tmp.name, data_only=True)
    except Exception as e:
        return jsonify({'error': f'Tiedoston avaus epäonnistui: {e}'}), 400

    recipes, errors = [], []
    for ws in wb.worksheets:
        try:
            # A sheet may contain one or many recipes ('Nimi' delimits each)
            name_rows = [r for r in range(1, ws.max_row + 1)
                         if str(ws.cell(r, 1).value or '').strip() == 'Nimi']
            if not name_rows:
                continue
            boundaries = name_rows + [ws.max_row + 1]
            for idx, start in enumerate(name_rows):
                end = boundaries[idx + 1]
                name = str(ws.cell(start, 2).value or '').strip()
                if not name:
                    continue
                diet, ingredients, ingredients_struct, instructions = '', [], [], []
                servings = None
                in_rivit = False
                for r in range(start, end):
                    label = str(ws.cell(r, 1).value or '').strip()
                    if label == 'Ruokavaliot':
                        diet = str(ws.cell(r, 2).value or '').strip()
                    if label == 'Annosmäärä':
                        try:
                            servings = int(ws.cell(r, 2).value)
                        except (TypeError, ValueError):
                            servings = None
                    if label == 'Rivit' or str(ws.cell(r, 2).value or '').strip() == 'Otsikko':
                        in_rivit = True
                        continue
                    if in_rivit:
                        ing = str(ws.cell(r, 3).value or '').strip()
                        # Ostopaino (sarake F) ensisijaisesti — se on tilausmäärä;
                        # Valmistamaton paino (sarake D) varalla
                        qty_raw = ws.cell(r, 6).value
                        if qty_raw in (None, ''):
                            qty_raw = ws.cell(r, 4).value
                        if ing:
                            qty, unit = _parse_qty(qty_raw)
                            if qty is not None:
                                ingredients.append(ing)
                                ingredients_struct.append({'name': ing, 'qty': qty, 'unit': unit})
                            else:
                                # "Raaka-aine / työohje" -sarake kantaa myös
                                # työohjeita — oikeissa PoweResta-tiedostoissa
                                # ne ovat rivejä joilla ei ole määrää (yksi
                                # ohjerivi per valmistusvaihe/ryhmä).
                                instructions.append(ing)
                ing_text = ' '.join(ingredients)
                recipes.append({
                    'name_fi': name,
                    'source_url': '',
                    'source_site': f'PoweResta-tuonti ({f.filename})',
                    'ingredients_raw': ingredients,
                    'ingredients_struct': ingredients_struct,
                    'instructions_raw': '\n'.join(instructions),
                    'season': guess_season(name, ing_text) or 'kaikki',
                    'meal_type': 'lounas',
                    'dish_category': guess_category(name, ing_text),
                    'prep_time_min': None,
                    'difficulty': None,
                    'servings': servings,
                    'notes': f'Ruokavaliot: {diet}' if diet else '',
                    'scraped_at': _dt.now().isoformat(),
                })
        except Exception as e:
            errors.append(f'{ws.title}: {e}')

    wb.close()
    try:
        os.unlink(tmp.name)
    except OSError:
        pass  # temp-tiedosto siivoutuu viimeistään uudelleenkäynnistyksessä

    if not recipes:
        return jsonify({'error': 'Tiedostosta ei löytynyt reseptejä '
                                 '(odotettu PoweResta-muoto: "Nimi"-rivit)'}), 400
    added = _queue_append(recipes)
    msg = f'{added} reseptiä lisätty tarkistusjonoon ({len(recipes) - added} ohitettu, jo jonossa)'
    if errors:
        msg += f' — {len(errors)} välilehteä ohitettu virheen takia'
    return jsonify({'message': msg, 'added': added})


@app.route('/api/import/paste', methods=['POST'])
def import_paste():
    """
    Parse a manually copied recipe text into a review-queue item.
    First non-empty line = name; lines that look like quantities = ingredients;
    the rest = instructions. No automation involved - the user copies the
    text themselves from wherever they are allowed to.
    """
    import re as _re
    from recipe_scraper_v2 import guess_category, guess_season
    from datetime import datetime as _dt

    text = (request.json or {}).get('text', '').strip()
    if not text:
        return jsonify({'error': 'Liitä reseptin teksti'}), 400

    lines = [l.strip() for l in text.splitlines()]
    lines = [l for l in lines if l]
    if not lines:
        return jsonify({'error': 'Tyhjä teksti'}), 400

    name = lines[0][:120]
    units = r'(kg|g|l|dl|cl|ml|rkl|tl|kpl|prk|pkt|ps|tlk|nippu|pala|viipale)'
    ing_pattern = _re.compile(r'^[\d¼½¾]|^n\.\s*\d|\d\s*' + units + r'\b', _re.IGNORECASE)
    ingredients, instructions = [], []
    for l in lines[1:]:
        if len(l) < 80 and ing_pattern.search(l):
            ingredients.append(l)
        else:
            instructions.append(l)

    ing_text = ' '.join(ingredients) or text
    item = {
        'name_fi': name,
        'source_url': '',
        'source_site': 'Käsin liitetty',
        'ingredients_raw': ingredients,
        'instructions_raw': ' '.join(instructions)[:2000],
        'season': guess_season(name, ing_text) or 'kaikki',
        'meal_type': 'lounas',
        'dish_category': guess_category(name, ing_text),
        'prep_time_min': None,
        'difficulty': None,
        'servings': None,
        'notes': '',
        'scraped_at': _dt.now().isoformat(),
    }
    added = _queue_append([item])
    if added:
        return jsonify({'message': f"'{name}' lisätty tarkistusjonoon "
                                   f"({len(ingredients)} raaka-ainetta tunnistettu)"})
    return jsonify({'error': f"'{name}' on jo tarkistusjonossa"}), 409



# ============================================================
# USER-CONTRIBUTED RECIPES: kitchen staff type recipes in directly;
# an admin approves each one before it enters the shared recipe database.
# ============================================================

def _ensure_user_recipe_tables(conn):
    conn.executescript("""
        CREATE TABLE IF NOT EXISTS user_recipes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name_fi TEXT UNIQUE NOT NULL,
            season TEXT NOT NULL,
            meal_type TEXT DEFAULT 'lounas',
            dish_category TEXT NOT NULL,
            instructions_raw TEXT,
            notes TEXT,
            created_by_user_id INTEGER,
            created_by_username TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            approved INTEGER DEFAULT 0,
            source TEXT DEFAULT 'user-contributed'
        );
        CREATE TABLE IF NOT EXISTS user_recipe_ingredients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_recipe_id INTEGER NOT NULL,
            ingredient_name TEXT NOT NULL,
            quantity REAL,
            unit TEXT,
            ingredient_instruction TEXT
        );
    """)
    _cols = [r[1] for r in conn.execute('PRAGMA table_info(user_recipes)').fetchall()]
    if 'recipe_type' not in _cols:
        conn.execute("ALTER TABLE user_recipes ADD COLUMN recipe_type TEXT DEFAULT 'pääruoka'")


def _fetch_user_recipe_ingredients(conn, user_recipe_id):
    rows = conn.execute(
        '''SELECT ingredient_name, quantity, unit, ingredient_instruction
           FROM user_recipe_ingredients WHERE user_recipe_id=?''',
        (user_recipe_id,)).fetchall()
    return [dict(r) for r in rows]


@app.route('/api/recipes/contribute', methods=['POST'])
def contribute_recipe():
    """Staff submit a recipe with structured ingredient rows; it waits in
    user_recipes/user_recipe_ingredients for admin approval."""
    d = request.json or {}
    name = (d.get('name_fi') or '').strip()
    season = d.get('season')
    category = d.get('dish_category')
    if not name or season not in SEASONS or category not in CATEGORIES:
        return jsonify({'error': 'Nimi, kausi ja kategoria vaaditaan'}), 400
    recipe_type = d.get('recipe_type') if d.get('recipe_type') in RECIPE_TYPES else 'pääruoka'

    rows = [row for row in (d.get('ingredients') or []) if (row.get('name') or '').strip()]

    conn = get_db()
    try:
        _ensure_user_recipe_tables(conn)
        try:
            cur = conn.execute(
                '''INSERT INTO user_recipes
                   (name_fi, season, meal_type, dish_category, recipe_type, instructions_raw, notes,
                    created_by_user_id, created_by_username)
                   VALUES (?,?,?,?,?,?,?,?,?)''',
                (name, season, d.get('meal_type', 'lounas'), category, recipe_type,
                 d.get('instructions_raw', ''), d.get('notes', ''),
                 None, 'admin'))
        except sqlite3.IntegrityError:
            return jsonify({'error': f"'{name}' odottaa jo hyväksyntää tai on jo olemassa"}), 409
        uid = cur.lastrowid
        for row in rows:
            qty = row.get('amount')
            try:
                qty = round(float(qty), 3) if qty not in (None, '') else None
            except (TypeError, ValueError):
                qty = None
            conn.execute(
                '''INSERT INTO user_recipe_ingredients
                   (user_recipe_id, ingredient_name, quantity, unit, ingredient_instruction)
                   VALUES (?,?,?,?,?)''',
                (uid, row['name'].strip(), qty, (row.get('unit') or '').strip() or None,
                 (row.get('instruction') or '').strip() or None))
        conn.commit()
    finally:
        conn.close()
    return jsonify({'message': f"'{name}' lisätty odottamaan hyväksyntää "
                               f"({len(rows)} raaka-ainetta)"})


@app.route('/api/recipes/pending')
def list_pending_recipes():
    conn = get_db()
    try:
        _ensure_user_recipe_tables(conn)
        rows = conn.execute('''SELECT * FROM user_recipes WHERE approved=0
                               ORDER BY created_at''').fetchall()
        out = []
        for r in rows:
            item = dict(r)
            item['ingredients'] = _fetch_user_recipe_ingredients(conn, r['id'])
            out.append(item)
        return jsonify(out)
    finally:
        conn.close()


@app.route('/api/recipes/pending/<int:uid>/approve', methods=['POST'])
def approve_pending_recipe(uid):
    """Move a pending user recipe into the shared recipes + recipe_ingredients tables."""
    d = request.json or {}
    conn = get_db()
    try:
        _ensure_user_recipe_tables(conn)
        row = conn.execute('SELECT * FROM user_recipes WHERE id=? AND approved=0',
                           (uid,)).fetchone()
        if not row:
            return jsonify({'error': 'Reseptiä ei löytynyt tai se on jo käsitelty'}), 404
        name = (d.get('name_fi') or row['name_fi']).strip()
        season = d.get('season') or row['season']
        category = d.get('dish_category') or row['dish_category']
        recipe_type = d.get('recipe_type') or row['recipe_type'] or 'pääruoka'
        if recipe_type not in RECIPE_TYPES:
            recipe_type = 'pääruoka'
        ingredient_rows = _fetch_user_recipe_ingredients(conn, uid)
        ingredients_struct = []
        for ing in ingredient_rows:
            ing_name = ing['ingredient_name']
            if ing['ingredient_instruction']:
                ing_name = f"{ing_name} ({ing['ingredient_instruction']})"
            ingredients_struct.append({'name': ing_name, 'qty': ing['quantity'],
                                       'unit': ing['unit'] or 'kg'})

        db = MealPlanDB(DB_PATH)
        rid = db.add_recipe(
            name_fi=name, season=season, meal_type=row['meal_type'],
            dish_category=category, recipe_type=recipe_type,
            notes=((row['notes'] or '') + ' [käyttäjän lisäämä]').strip(),
            instructions=row['instructions_raw'] or '')
        if rid is None:
            existing = conn.execute('SELECT id FROM recipes WHERE name_fi=?', (name,)).fetchone()
            rid = existing['id'] if existing else None
        else:
            _ensure_recipe_source_column(conn)
            _set_recipe_source(conn, rid, 'user_contributed')
        _save_recipe_ingredients(rid, [it for it in ingredients_struct if it['qty'] is not None])
        # jätä rivi tauluun approved=1:llä (historia + PoweResta-vientiä varten)
        conn.execute('''UPDATE user_recipes SET name_fi=?, season=?, dish_category=?,
                        recipe_type=?, approved=1 WHERE id=?''',
                     (name, season, category, recipe_type, uid))
        conn.commit()
        return jsonify({'message': f"'{name}' hyväksytty ja lisätty reseptitietokantaan",
                        'recipe_id': rid})
    finally:
        conn.close()


@app.route('/api/recipes/pending/<int:uid>/reject', methods=['POST'])
def reject_pending_recipe(uid):
    conn = get_db()
    try:
        _ensure_user_recipe_tables(conn)
        conn.execute('DELETE FROM user_recipe_ingredients WHERE user_recipe_id=?', (uid,))
        conn.execute('DELETE FROM user_recipes WHERE id=?', (uid,))
        conn.commit()
        return jsonify({'message': 'Resepti hylätty'})
    finally:
        conn.close()


def _fmt_qty_unit(quantity, unit):
    """'2,5 kg' -muotoinen yhdistelmä PoweResta-tuontia varten (suomalainen desimaalipilkku)."""
    parts = []
    if quantity is not None:
        s = f'{quantity:.3f}'.rstrip('0').rstrip('.').replace('.', ',')
        parts.append(s)
    if unit:
        parts.append(unit)
    return ' '.join(parts)


@app.route('/api/recipes/export/poweresta')
def export_user_recipes_poweresta():
    """Vie hyväksytyt käyttäjän lisäämät reseptit PoweResta-muotoisena Excelinä
    (varmuuskopiointiin, jakoon toisille keittiöille tai arkistointiin)."""
    from openpyxl import Workbook

    conn = get_db()
    try:
        _ensure_user_recipe_tables(conn)
        recipes = conn.execute(
            '''SELECT * FROM user_recipes WHERE approved=1
               ORDER BY season, name_fi''').fetchall()
        if not recipes:
            return jsonify({'error': 'Hyväksyttyjä reseptejä ei löytynyt'}), 404

        wb = Workbook()
        wb.remove(wb.active)  # tyhjä oletusvälilehti pois
        used_names = set()

        for r in recipes:
            base = (r['name_fi'] or 'Resepti')
            for ch in '[]:*?/\\':
                base = base.replace(ch, ' ')
            base = base.strip()[:31] or 'Resepti'
            sheet_name = base
            n = 2
            while sheet_name in used_names:
                suffix = f' ({n})'
                sheet_name = base[:31 - len(suffix)] + suffix
                n += 1
            used_names.add(sheet_name)
            ws = wb.create_sheet(sheet_name)

            ws.cell(1, 1, 'Nimi')
            ws.cell(1, 2, r['name_fi'])
            ws.cell(2, 1, 'Ruokavaliot')
            ws.cell(2, 2, r['season'])
            ws.cell(3, 1, 'Rivit')

            ingredients = conn.execute(
                '''SELECT ingredient_name, quantity, unit, ingredient_instruction
                   FROM user_recipe_ingredients WHERE user_recipe_id=?''',
                (r['id'],)).fetchall()

            row = 4
            for ing in ingredients:
                name = ing['ingredient_name']
                if ing['ingredient_instruction']:
                    name = f"{name} ({ing['ingredient_instruction']})"
                ws.cell(row, 3, name)
                ws.cell(row, 4, _fmt_qty_unit(ing['quantity'], ing['unit']))
                row += 1

            if r['instructions_raw']:
                ws.cell(row, 3, r['instructions_raw'])

        stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        out = os.path.join(OUTPUT_DIR, f'user_recipes_poweresta_{stamp}.xlsx')
        wb.save(out)
        return send_file(out, as_attachment=True,
                         download_name=f'user_recipes_poweresta_{stamp}.xlsx')
    finally:
        conn.close()


_DIET_NOTE_RE_STR = r'Ruokavaliot:\s*([^\n]*)'


@app.route('/api/recipes/export-selected/poweresta')
def export_selected_recipes_poweresta():
    """Vie valitut (tai kaikki) reseptit PoweResta-muotoisena Excelinä —
    minkä tahansa reseptin, ei vain käyttäjän lisäämien (ks. myös
    /api/recipes/export/poweresta, joka on eri, admin-rajattu ominaisuus
    käyttäjän lisäämille resepteille)."""
    import re as _re
    from openpyxl import Workbook

    recipe_ids_str = request.args.get('recipe_ids', '')
    export_all = request.args.get('all', 'false').lower() == 'true'

    conn = get_db()
    try:
        _ensure_ingredient_editor_tables(conn)
        if export_all:
            recipes = conn.execute('SELECT * FROM recipes ORDER BY name_fi').fetchall()
        elif recipe_ids_str:
            try:
                ids = [int(x.strip()) for x in recipe_ids_str.split(',') if x.strip()]
            except ValueError:
                return jsonify({'error': 'recipe_ids pitää olla pilkulla erotettu lista numeroita'}), 400
            if not ids:
                return jsonify({'error': 'Valitse reseptit (recipe_ids tai all=true)'}), 400
            placeholders = ','.join('?' * len(ids))
            recipes = conn.execute(
                f'SELECT * FROM recipes WHERE id IN ({placeholders}) ORDER BY name_fi', ids).fetchall()
        else:
            return jsonify({'error': 'Valitse reseptit (recipe_ids tai all=true)'}), 400

        if not recipes:
            return jsonify({'error': 'Reseptejä ei löytynyt'}), 404

        wb = Workbook()
        wb.remove(wb.active)
        used_names = set()

        for r in recipes:
            base = (r['name_fi'] or 'Resepti')
            for ch in '[]:*?/\\':
                base = base.replace(ch, ' ')
            base = base.strip()[:31] or 'Resepti'
            sheet_name = base
            n = 2
            while sheet_name in used_names:
                suffix = f' ({n})'
                sheet_name = base[:31 - len(suffix)] + suffix
                n += 1
            used_names.add(sheet_name)
            ws = wb.create_sheet(sheet_name)

            # Ruokavaliot: käytä oikeaa dieettikoodia jos se on tallennettu
            # notes-kenttään PoweResta-tuonnista ("Ruokavaliot: GL,LA,..."),
            # muuten jätä tyhjäksi — sesonkia EI laiteta tähän, koska se ei
            # ole dieettitieto eikä sitä pidä esittää sellaisena.
            diet = ''
            if r['notes']:
                m = _re.search(_DIET_NOTE_RE_STR, r['notes'])
                if m:
                    diet = m.group(1).strip()

            ws.cell(1, 1, 'Nimi')
            ws.cell(1, 2, r['name_fi'])
            ws.cell(2, 1, 'Ruokavaliot')
            ws.cell(2, 2, diet)
            ws.cell(3, 1, 'Rivit')
            ws.cell(3, 3, 'Raaka-aine / työohje')
            ws.cell(3, 4, 'Määrä')

            ingredients = conn.execute(
                '''SELECT i.name_fi AS ingredient_name, ri.quantity, ri.unit, ri.ingredient_instruction
                   FROM recipe_ingredients ri JOIN ingredients i ON i.id = ri.ingredient_id
                   WHERE ri.recipe_id=? ORDER BY ri.sort_order, i.name_fi''', (r['id'],)).fetchall()

            row = 4
            for ing in ingredients:
                name = ing['ingredient_name']
                if ing['ingredient_instruction']:
                    name = f"{name} ({ing['ingredient_instruction']})"
                ws.cell(row, 3, name)
                ws.cell(row, 4, _fmt_qty_unit(ing['quantity'], ing['unit']))
                row += 1

            if not ingredients:
                ws.cell(row, 3, '(Ei raaka-ainetietoja tässä järjestelmässä)')
                row += 1

            # Muut tiedot (huom: ei ole erillistä valmistusohje-kenttää tässä
            # järjestelmässä — notes voi sisältää sekalaisia lisätietoja)
            other_notes = r['notes']
            if other_notes:
                other_notes = _re.sub(_DIET_NOTE_RE_STR, '', other_notes).strip()
            if other_notes:
                ws.cell(row + 1, 3, f'Muut tiedot: {other_notes}')

            for col, width in zip('ACD', (16, 45, 16)):
                ws.column_dimensions[col].width = width

        stamp = datetime.now().strftime('%Y%m%d')
        out_name = f'reseptit_poweresta_{stamp}.xlsx'
        out = os.path.join(OUTPUT_DIR, out_name)
        wb.save(out)
        return send_file(out, as_attachment=True, download_name=out_name)
    finally:
        conn.close()


# ============================================================
# OCR IMPORT: scanned/photographed weekly menu tables
# ============================================================

ocr_state = {'running': False, 'total': 0, 'done': 0, 'ok': 0, 'failed': 0,
             'dishes': 0, 'log': [], 'error': None}
ocr_lock = threading.Lock()


def _ocr_worker(paths_names, main_meals_only):
    from ocr_menu_import import extract_week_menu, menu_to_queue_items
    try:
        for path, name in paths_names:
            try:
                parsed = extract_week_menu(path)
                items = menu_to_queue_items(parsed, name, main_meals_only)
                added = _queue_append(items)
                with ocr_lock:
                    ocr_state['done'] += 1
                    ocr_state['ok'] += 1
                    ocr_state['dishes'] += added
                    wk = parsed.get('week_label') or '?'
                    ocr_state['log'].append(
                        f'✅ {name} ({wk}): {added} uutta ruokaa jonoon')
            except Exception as e:
                with ocr_lock:
                    ocr_state['done'] += 1
                    ocr_state['failed'] += 1
                    ocr_state['log'].append(f'❌ {name}: {e}')
            finally:
                try:
                    os.unlink(path)
                except OSError:
                    pass
    finally:
        with ocr_lock:
            ocr_state['running'] = False


@app.route('/api/import/ocr', methods=['POST'])
def import_ocr():
    """Upload photos/scans of weekly menu tables; extract dishes via OCR."""
    with ocr_lock:
        if ocr_state['running']:
            return jsonify({'error': 'OCR-käsittely on jo käynnissä'}), 409

    files = request.files.getlist('files')
    files = [f for f in files if f.filename]
    if not files:
        return jsonify({'error': 'Valitse vähintään yksi kuva'}), 400

    # Fail fast with clear guidance if Tesseract isn't installed
    from ocr_menu_import import _tesseract_available
    if not _tesseract_available():
        return jsonify({'error': 'Tesseract OCR (fin) puuttuu koneelta. '
                                 'Katso asennusohje OHJEET.md:stä kohdasta OCR.'}), 500

    import tempfile
    paths = []
    for f in files[:60]:
        ext = os.path.splitext(f.filename)[1].lower() or '.jpg'
        tmp = tempfile.NamedTemporaryFile(suffix=ext, delete=False)
        f.save(tmp.name)
        paths.append((tmp.name, f.filename))

    main_only = request.form.get('main_meals_only', '1') == '1'
    with ocr_lock:
        ocr_state.update({'running': True, 'total': len(paths), 'done': 0,
                          'ok': 0, 'failed': 0, 'dishes': 0, 'log': [], 'error': None})
    threading.Thread(target=_ocr_worker, args=(paths, main_only), daemon=True).start()
    return jsonify({'message': f'OCR aloitettu: {len(paths)} kuvaa', 'total': len(paths)})


@app.route('/api/import/ocr/status')
def ocr_status():
    with ocr_lock:
        return jsonify(dict(ocr_state))


# ============================================================
# SCRAPING: fetch recipes from the web into the review queue
# ============================================================

scrape_state = {
    'running': False, 'total': 0, 'done': 0, 'ok': 0, 'failed': 0,
    'log': [], 'error': None
}
scrape_lock = threading.Lock()


def _scrape_worker(urls):
    """Background worker: scrape each URL, append results to review queue."""
    from recipe_scraper_v2 import scrape_recipe
    import time as _time

    try:
        try:
            with open(REVIEW_FILE, encoding='utf-8') as f:
                queue = json.load(f)
        except (FileNotFoundError, json.JSONDecodeError):
            queue = []
        existing = {r.get('source_url') for r in queue}

        for url in urls:
            with scrape_lock:
                if url in existing:
                    scrape_state['done'] += 1
                    scrape_state['log'].append(f'⏭️ Jo jonossa: {url}')
                    continue
            try:
                recipe = scrape_recipe(url)
                with scrape_lock:
                    scrape_state['done'] += 1
                    if recipe and recipe.get('name_fi'):
                        queue.append(recipe)
                        existing.add(url)
                        scrape_state['ok'] += 1
                        cat = recipe.get('dish_category') or '?'
                        sea = recipe.get('season') or 'ympärivuotinen?'
                        scrape_state['log'].append(
                            f"✅ {recipe['name_fi']} (kategoria: {cat}, kausi: {sea})")
                    else:
                        scrape_state['failed'] += 1
                        scrape_state['log'].append(
                            f'⚠️ Ei reseptimerkintää sivulla: {url}')
            except Exception as e:
                with scrape_lock:
                    scrape_state['done'] += 1
                    scrape_state['failed'] += 1
                    scrape_state['log'].append(f'❌ {url}: {e}')
            _time.sleep(1.0)  # be polite to the source site

        with open(REVIEW_FILE, 'w', encoding='utf-8') as f:
            json.dump(queue, f, ensure_ascii=False, indent=2)
    except Exception as e:
        with scrape_lock:
            scrape_state['error'] = str(e)
    finally:
        with scrape_lock:
            scrape_state['running'] = False


@app.route('/api/scrape', methods=['POST'])
def start_scrape():
    """Start scraping. Body: {urls: 'one per line'} or {listing_url, limit}."""
    with scrape_lock:
        if scrape_state['running']:
            return jsonify({'error': 'Haku on jo käynnissä — odota sen valmistumista'}), 409

    data = request.json or {}
    urls = []

    if data.get('listing_url'):
        from recipe_scraper_v2 import collect_recipe_links
        try:
            limit = min(int(data.get('limit', 20)), 50)
            urls = collect_recipe_links(data['listing_url'].strip(), limit=limit)
        except Exception as e:
            return jsonify({'error': f'Listaussivun haku epäonnistui: {e}'}), 400
        if not urls:
            return jsonify({'error': 'Sivulta ei löytynyt reseptilinkkejä. '
                                     'Kokeile liittää reseptisivujen osoitteet suoraan.'}), 400
    else:
        raw = data.get('urls', '')
        urls = [u.strip() for u in raw.splitlines() if u.strip().startswith('http')]
        if not urls:
            return jsonify({'error': 'Liitä vähintään yksi osoite (http...)'}), 400
        urls = urls[:50]

    with scrape_lock:
        scrape_state.update({'running': True, 'total': len(urls), 'done': 0,
                             'ok': 0, 'failed': 0, 'log': [], 'error': None})

    threading.Thread(target=_scrape_worker, args=(urls,), daemon=True).start()
    return jsonify({'message': f'Haku aloitettu: {len(urls)} osoitetta', 'total': len(urls)})


@app.route('/api/scrape/status')
def scrape_status():
    with scrape_lock:
        return jsonify(dict(scrape_state))


# ============================================================
# STATS
# ============================================================

def _latest_backup_timestamp():
    """Uusimman varmuuskopion aikaleima ('varmuuskopiot/'-kansiosta), tai None."""
    import re as _re
    backup_dir = os.path.join(os.path.dirname(DB_PATH), 'varmuuskopiot')
    name_re = _re.compile(r'^meal_plans_(\d{8})_(\d{6})\.db$')
    if not os.path.isdir(backup_dir):
        return None
    names = sorted(f for f in os.listdir(backup_dir) if name_re.match(f))
    if not names:
        return None
    m = name_re.match(names[-1])
    date_part, time_part = m.group(1), m.group(2)
    return (f'{date_part[6:8]}.{date_part[4:6]}.{date_part[0:4]} '
            f'{time_part[0:2]}:{time_part[2:4]}')


@app.route('/api/stats')
def stats():
    conn = get_db()
    by_season = {s: conn.execute(
        'SELECT COUNT(*) FROM recipes WHERE season = ?', (s,)).fetchone()[0]
        for s in SEASONS}
    by_category = {c: conn.execute(
        'SELECT COUNT(*) FROM recipes WHERE dish_category = ?', (c,)).fetchone()[0]
        for c in CATEGORIES}
    plans = conn.execute('SELECT COUNT(*) FROM meal_plans').fetchone()[0]
    recipes_total = conn.execute('SELECT COUNT(*) FROM recipes').fetchone()[0]
    meal_slots_total = conn.execute('SELECT COUNT(*) FROM meal_plan_days').fetchone()[0]
    users_total = conn.execute('SELECT COUNT(*) FROM users').fetchone()[0]
    conn.close()
    return jsonify({'by_season': by_season, 'by_category': by_category, 'plans': plans,
                    'recipes_total': recipes_total, 'meal_slots_total': meal_slots_total,
                    'users_total': users_total, 'last_backup': _latest_backup_timestamp()})


# ============================================================
# HEALTH: järjestelmän terveystarkistus (admin only)
# ============================================================

BACKUP_NAME_RE_STR = r'^meal_plans_(\d{8})_(\d{6})\.db$'
USER_CONTRIBUTED_MARK = '[käyttäjän lisäämä]'
NIGHTLY_BACKUP_HOUR = 22  # ks. nightly_backup.bat / Windowsin Task Scheduler -tehtävä


def _backup_dir():
    return os.path.join(os.path.dirname(DB_PATH), 'varmuuskopiot')


def _list_backups():
    """Kaikki varmuuskopiot aikaleimoineen ja koroineen, uusin ensin.
    Sisältää sekä käsin luodut että Task Schedulerin tekemät kopiot —
    ne jakavat saman kansion ja nimeämiskäytännön."""
    import re as _re
    backup_dir = _backup_dir()
    name_re = _re.compile(BACKUP_NAME_RE_STR)
    if not os.path.isdir(backup_dir):
        return []
    items = []
    for f in os.listdir(backup_dir):
        m = name_re.match(f)
        if not m:
            continue
        p = os.path.join(backup_dir, f)
        ts = datetime.strptime(m.group(1) + m.group(2), '%Y%m%d%H%M%S')
        items.append({'name': f, 'timestamp': ts, 'size_bytes': os.path.getsize(p)})
    items.sort(key=lambda b: b['timestamp'], reverse=True)
    return items


def _next_nightly_backup_time(after):
    """Seuraava ajankohta jolloin Task Schedulerin pitäisi ajaa varmuuskopio
    (staattinen laskelma kiinteästä klo 22:00 -aikataulusta — ei kysytä
    Windowsin Task Scheduleria suoraan, ks. nightly_backup.bat)."""
    today_22 = after.replace(hour=NIGHTLY_BACKUP_HOUR, minute=0, second=0, microsecond=0)
    return today_22 if after < today_22 else today_22 + timedelta(days=1)


@app.route('/api/health/status')
def health_status():
    backups = _list_backups()
    alerts = []
    severity = 'ok'
    if not backups:
        alerts.append('Yhtään varmuuskopiota ei ole vielä luotu')
        severity = 'critical'
    else:
        age_hours = (datetime.now() - backups[0]['timestamp']).total_seconds() / 3600
        if age_hours > 7 * 24:
            alerts.append('Viimeisin varmuuskopio on yli 7 päivää vanha')
            severity = 'critical'
        elif age_hours > 48:
            alerts.append('Viimeisin varmuuskopio on yli 48 tuntia vanha')
            severity = 'warning'

    return jsonify({
        'status': severity,
        'uptime_minutes': int((time.time() - _APP_START_TIME) / 60),
        'last_check': datetime.now().isoformat(),
        'alerts': alerts,
    })


@app.route('/api/health/database')
def health_database():
    conn = get_db()
    recipes_total = conn.execute('SELECT COUNT(*) FROM recipes').fetchone()[0]
    recipes_user_contributed = conn.execute(
        'SELECT COUNT(*) FROM recipes WHERE notes LIKE ?',
        (f'%{USER_CONTRIBUTED_MARK}%',)).fetchone()[0]
    recipes_without_ingredients = conn.execute(
        '''SELECT COUNT(*) FROM recipes r WHERE NOT EXISTS
           (SELECT 1 FROM recipe_ingredients ri WHERE ri.recipe_id = r.id)''').fetchone()[0]
    meal_plans_total = conn.execute('SELECT COUNT(*) FROM meal_plans').fetchone()[0]
    order_catalogs_total = conn.execute(
        "SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='order_catalogs'").fetchone()[0]
    if order_catalogs_total:
        order_catalogs_total = conn.execute('SELECT COUNT(*) FROM order_catalogs').fetchone()[0]
    users_total = conn.execute('SELECT COUNT(*) FROM users').fetchone()[0]
    last_modified = conn.execute('SELECT MAX(created_at) FROM recipes').fetchone()[0]
    conn.close()

    return jsonify({
        'size_mb': round(os.path.getsize(DB_PATH) / (1024 * 1024), 2),
        'recipes_total': recipes_total,
        'recipes_from_poweresta': recipes_total - recipes_user_contributed,
        'recipes_user_contributed': recipes_user_contributed,
        'recipes_without_ingredients': recipes_without_ingredients,
        'meal_plans_total': meal_plans_total,
        'order_catalogs_total': order_catalogs_total,
        'users_total': users_total,
        'last_modified': last_modified,
    })


@app.route('/api/health/backups')
def health_backups():
    backups = _list_backups()
    now = datetime.now()
    if not backups:
        return jsonify({
            'latest_backup': None, 'backup_count': 0, 'total_backup_size_mb': 0,
            'hours_since_last_backup': None,
            'next_automated_backup': _next_nightly_backup_time(now).isoformat(),
            'backup_schedule': f'Päivittäin klo {NIGHTLY_BACKUP_HOUR:02d}:00 (Windows Task Scheduler)',
        })
    latest = backups[0]
    total_bytes = sum(b['size_bytes'] for b in backups)
    return jsonify({
        'latest_backup': {'name': latest['name'], 'timestamp': latest['timestamp'].isoformat(),
                          'size_mb': round(latest['size_bytes'] / (1024 * 1024), 2)},
        'backup_count': len(backups),
        'total_backup_size_mb': round(total_bytes / (1024 * 1024), 2),
        'hours_since_last_backup': round((now - latest['timestamp']).total_seconds() / 3600, 1),
        'next_automated_backup': _next_nightly_backup_time(now).isoformat(),
        'backup_schedule': f'Päivittäin klo {NIGHTLY_BACKUP_HOUR:02d}:00 (Windows Task Scheduler)',
    })


@app.route('/api/health/quality')
def health_quality():
    conn = get_db()
    recipes_without_ingredients = conn.execute(
        '''SELECT r.id, r.name_fi FROM recipes r WHERE NOT EXISTS
           (SELECT 1 FROM recipe_ingredients ri WHERE ri.recipe_id = r.id)
           ORDER BY r.name_fi''').fetchall()
    empty_meal_plans = conn.execute(
        '''SELECT mp.id, mp.name FROM meal_plans mp WHERE NOT EXISTS
           (SELECT 1 FROM meal_plan_days d WHERE d.meal_plan_id = mp.id)
           ORDER BY mp.name''').fetchall()
    conn.close()

    warnings = []
    if recipes_without_ingredients:
        warnings.append({
            'type': 'recipes_missing_ingredients',
            'count': len(recipes_without_ingredients),
            'severity': 'medium',
            'description': f'{len(recipes_without_ingredients)} reseptiä ilman raaka-ainetietoja '
                           f'(yleensä OCR-tuonnista tai käsin lisätystä reseptistä ilman määriä)',
        })
    if empty_meal_plans:
        warnings.append({
            'type': 'empty_meal_plans',
            'count': len(empty_meal_plans),
            'severity': 'low',
            'description': f'{len(empty_meal_plans)} ruokalistaa ilman yhtään ateriaa',
        })

    return jsonify({
        'recipes_without_ingredients': len(recipes_without_ingredients),
        'recipes_without_ingredients_sample': [dict(r) for r in recipes_without_ingredients[:50]],
        'empty_meal_plans': len(empty_meal_plans),
        'warnings': warnings,
    })


@app.route('/api/health/integrity-check', methods=['POST'])
def health_integrity_check():
    start = time.time()
    conn = get_db()
    errors = []

    # Raakatason SQLite-tiedostotarkistus (btree/sivukorruptio) — eri asia kuin
    # sovellustason orpo-rivien tarkistus alla, jota tämä ei korvaa.
    integrity = conn.execute('PRAGMA integrity_check').fetchone()[0]
    if integrity != 'ok':
        errors.append(f'SQLite-tiedoston eheystarkistus epäonnistui: {integrity}')

    bad_recipe_types = conn.execute(
        "SELECT COUNT(*) FROM recipes WHERE recipe_type NOT IN "
        "('pääruoka','keitto','salaatti','leivonta','jälkiruoka',"
        "'kastike','kasvislisäke','energialisäke')"
    ).fetchone()[0]
    if bad_recipe_types:
        errors.append(f'{bad_recipe_types} reseptillä on virheellinen recipe_type-arvo')

    bad_roles = conn.execute(
        "SELECT COUNT(*) FROM users WHERE role NOT IN ('admin','muokkaus','katselu')"
    ).fetchone()[0]
    if bad_roles:
        errors.append(f'{bad_roles} käyttäjällä on virheellinen rooli')

    orphan_ri = conn.execute(
        '''SELECT COUNT(*) FROM recipe_ingredients
           WHERE recipe_id NOT IN (SELECT id FROM recipes)''').fetchone()[0]
    if orphan_ri:
        errors.append(f'{orphan_ri} orpoa recipe_ingredients-riviä (resepti poistettu)')

    orphan_days = conn.execute(
        '''SELECT COUNT(*) FROM meal_plan_days
           WHERE meal_plan_id NOT IN (SELECT id FROM meal_plans)
              OR recipe_id NOT IN (SELECT id FROM recipes)''').fetchone()[0]
    if orphan_days:
        errors.append(f'{orphan_days} orpoa meal_plan_days-riviä (ruokalista tai resepti puuttuu)')

    has_catalog_items = conn.execute(
        "SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='order_catalog_items'").fetchone()[0]
    if has_catalog_items:
        orphan_items = conn.execute(
            '''SELECT COUNT(*) FROM order_catalog_items
               WHERE catalog_id NOT IN (SELECT id FROM order_catalogs)''').fetchone()[0]
        if orphan_items:
            errors.append(f'{orphan_items} orpoa tilauslistarivi (tilauslista puuttuu)')

    has_user_recipe_ing = conn.execute(
        "SELECT COUNT(*) FROM sqlite_master WHERE type='table' AND name='user_recipe_ingredients'").fetchone()[0]
    if has_user_recipe_ing:
        orphan_uri = conn.execute(
            '''SELECT COUNT(*) FROM user_recipe_ingredients
               WHERE user_recipe_id NOT IN (SELECT id FROM user_recipes)''').fetchone()[0]
        if orphan_uri:
            errors.append(f'{orphan_uri} orpoa user_recipe_ingredients-riviä')

    conn.close()
    checked = ['recipes', 'recipe_ingredients', 'meal_plans', 'meal_plan_days', 'users']
    if has_catalog_items:
        checked += ['order_catalogs', 'order_catalog_items']
    if has_user_recipe_ing:
        checked += ['user_recipes', 'user_recipe_ingredients']

    return jsonify({
        'status': 'ok' if not errors else 'error',
        'errors_found': len(errors),
        'errors': errors,
        'scan_time_seconds': round(time.time() - start, 2),
        'checked_tables': checked,
    })


# requirements.txt -> (import-nimi, ihmisluettava nimi). pymupdf tuodaan
# nimellä 'fitz', ei 'pymupdf' — ks. menu_pdf_generator.py.
_DEPENDENCY_MODULES = [
    ('flask', 'Flask'), ('werkzeug', 'Werkzeug'), ('openpyxl', 'openpyxl'),
    ('requests', 'requests'), ('bs4', 'beautifulsoup4'), ('docx', 'python-docx'),
    ('fitz', 'PyMuPDF'), ('reportlab', 'reportlab'), ('PIL', 'Pillow'),
    ('pytesseract', 'pytesseract'), ('ctranslate2', 'ctranslate2'),
    ('sentencepiece', 'sentencepiece'),
]


@app.route('/api/health/dependencies')
def health_dependencies():
    import importlib
    results = []
    for module_name, label in _DEPENDENCY_MODULES:
        try:
            importlib.import_module(module_name)
            results.append({'name': label, 'ok': True})
        except ImportError as e:
            results.append({'name': label, 'ok': False, 'error': str(e)})

    model_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'models', 'translate-en_fi')
    translation_model_present = os.path.exists(os.path.join(model_dir, 'sentencepiece.model'))
    results.append({'name': 'Käännösmalli (en->fi)', 'ok': translation_model_present})

    missing = [r['name'] for r in results if not r['ok']]
    return jsonify({
        'status': 'ok' if not missing else 'error',
        'checked': len(results),
        'missing': missing,
        'results': results,
    })


if __name__ == '__main__':
    MealPlanDB(DB_PATH)  # ensure tables exist
    MealModifier(DB_PATH)  # ensure modification tables exist
    print('=' * 60)
    print('  Ruokalistasuunnittelija käynnissä')
    print('  Avaa selaimessa: http://localhost:5001')
    print('=' * 60)
    app.run(debug=False, host='0.0.0.0', port=5001)
